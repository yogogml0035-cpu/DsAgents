from __future__ import annotations

import copy
import math
import re
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping, Sequence
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from workflow_artifacts import (
    read_json_artifact,
    resolve_artifact_path,
    to_virtual_artifact_path,
    unique_download_path,
    write_json_artifact,
)


WORKFLOW = "tecan-import"
EXTRACTORS = frozenset(f"tecan-extractor-{name}" for name in "abc")
CONFIDENCES = frozenset({"high", "medium", "low"})
LOGISTICS_FIELDS = ("pieces", "gross_weight")
CANONICAL_KEYS = frozenset(
    {"workflow", "source_artifacts", "logistics", "items", "manual_checks", "currency"}
)
ITEM_KEYS = frozenset(
    {
        "pn",
        "quantity",
        "net_price",
        "amount",
        "description",
        "origin_country",
        "unit_net_weight",
        "net_weight",
        "gross_weight",
        "source_artifact",
        "source_sheet",
        "source_row",
    }
)
_TEMPLATE = Path(__file__).resolve().parent / "skills" / WORKFLOW / "assets" / "Tecan_进口_发票箱单_空运.xlsx"
_WEIGHT_QUANT = Decimal("0.001")
_CURRENCY_RE = re.compile(r"US\$|RMB|CNY|USD|EUR|GBP|JPY|SGD|HKD|CHF|AUD|CAD|[¥￥$€£]", re.IGNORECASE)


def save_tecan_extraction(
    extractor: str,
    source_artifact: str,
    logistics: dict[str, dict[str, Any]],
    items: list[dict[str, Any]],
) -> dict[str, str]:
    """Validate and immutably save one Tecan logistics extraction artifact."""
    if not resolve_artifact_path(source_artifact).is_file():
        raise ValueError(f"source_artifact not found: {source_artifact}")
    payload = _validate_extraction(
        {
            "workflow": WORKFLOW,
            "extractor": extractor,
            "source_artifact": source_artifact,
            "logistics": logistics,
            "items": items,
        }
    )
    path = write_json_artifact(extractor, payload)
    return {"extractor": extractor, "artifact_path": path}


def save_tecan_adjudication(
    source_artifacts: list[str],
    decisions: list[dict[str, Any]],
) -> dict[str, str]:
    """Save a small Tecan logistics adjudication artifact."""
    if not source_artifacts:
        raise ValueError("source_artifacts must not be empty")
    for path in source_artifacts:
        if not resolve_artifact_path(path).is_file():
            raise ValueError(f"source_artifact not found: {path}")
    path = write_json_artifact(
        "tecan_adjudication",
        {"workflow": WORKFLOW, "source_artifacts": source_artifacts, "decisions": _validate_decisions(decisions)},
    )
    return {"artifact_path": path}


def build_tecan_canonical(
    extraction_artifacts: list[str],
    order_artifact: str,
    information_artifacts: list[str],
    info_source_preference: str | None = None,
    pn_info_source_overrides: dict[str, str] | None = None,
    adjudication_artifact: str | None = None,
) -> dict[str, Any]:
    """Reconcile logistics and deterministically join explicit order/information workbooks."""
    order_path = resolve_artifact_path(order_artifact)
    info_paths = [(path, resolve_artifact_path(path)) for path in information_artifacts]
    if not order_path.is_file() or not info_paths:
        return {"status": "needs_input", "missing": ["order_or_information_artifact"]}

    payloads, invalid = _load_extractions(extraction_artifacts)
    if not payloads:
        return {"status": "needs_input", "missing": ["valid_extraction"]}
    extractor_ids = {payload["extractor"] for payload in payloads}
    if len(extractor_ids) != len(payloads):
        raise ValueError("Duplicate extractor artifacts are not allowed")
    source_artifacts = {payload["source_artifact"] for payload in payloads}
    if len(source_artifacts) != 1:
        raise ValueError("All extraction artifacts must reference the same source_artifact")
    source_artifact = next(iter(source_artifacts))
    initial_ab = extractor_ids <= {"tecan-extractor-a", "tecan-extractor-b"}
    if initial_ab and adjudication_artifact is not None:
        raise ValueError("Extractor C is required before adjudication")
    if initial_ab and len(payloads) < 2:
        return _needs_c(source_artifact, ["only_one_extractor_succeeded", *invalid])

    adjudication = _load_adjudication(adjudication_artifact) if adjudication_artifact else None
    if adjudication and adjudication["source_artifacts"] != extraction_artifacts:
        raise ValueError("Adjudication source_artifacts do not match extraction_artifacts")
    decisions = {item["conflict_id"]: item["value"] for item in (adjudication or {}).get("decisions", [])}
    used_decisions: set[str] = set()
    conflicts: list[dict[str, Any]] = []
    manual_checks: list[dict[str, Any]] = []
    missing: list[str] = []
    logistics: dict[str, Any] = {}
    for field_name in LOGISTICS_FIELDS:
        value = _resolve_vote(
            f"logistics.{field_name}",
            field_name,
            [(payload["extractor"], payload["logistics"][field_name]) for payload in payloads],
            decisions,
            used_decisions,
            conflicts,
            manual_checks,
        )
        logistics[field_name] = value
        if value is None:
            missing.append(f"logistics.{field_name}")

    if initial_ab and (conflicts or missing):
        return _needs_c(source_artifact, [item["field"] for item in conflicts] + missing)
    if conflicts:
        unknown = sorted(set(decisions) - used_decisions)
        if unknown:
            raise ValueError("Unknown conflict_id: " + ", ".join(unknown))
        return {"status": "needs_adjudication", "source_artifact": source_artifact, "conflicts": conflicts}
    if set(decisions) - used_decisions:
        raise ValueError("Adjudication contains conflict_id values that are not present")
    if missing:
        return {"status": "needs_input", "missing": missing}

    order = _parse_order_workbook(order_path, order_artifact)
    if order.get("status") != "ok":
        return order
    info_records = _parse_information_workbooks(info_paths)
    if isinstance(info_records, dict) and info_records.get("status") == "needs_input":
        return info_records

    candidates_by_pn = info_records
    assert isinstance(candidates_by_pn, dict)
    overrides = {normalize_pn(pn): choice for pn, choice in (pn_info_source_overrides or {}).items()}
    selected: dict[str, dict[str, Any]] = {}
    info_conflicts: list[dict[str, Any]] = []
    for row in order["rows"]:
        pn = row["pn"]
        candidates = candidates_by_pn.get(pn, [])
        if not candidates:
            return {"status": "needs_input", "missing": [f"information.{pn}"]}
        distinct = _distinct_info_records(candidates)
        if len(distinct) == 1:
            selected[pn] = distinct[0]
            continue
        choice = overrides.get(pn, info_source_preference)
        picked = _select_info_record(distinct, choice)
        if picked is None:
            info_conflicts.append(
                {
                    "pn": pn,
                    "candidates": [
                        {
                            "source_artifact": item["source_artifact"],
                            "source_sheet": item["source_sheet"],
                            "source_row": item["source_row"],
                            "description": item["description"],
                            "origin_country": item["origin_country"],
                            "unit_net_weight": item["unit_net_weight"],
                        }
                        for item in distinct
                    ],
                }
            )
        else:
            selected[pn] = picked
    if info_conflicts:
        return {"status": "needs_input", "reason": "information_conflict", "conflicts": info_conflicts}

    total_net = Decimal("0")
    items: list[dict[str, Any]] = []
    for row in order["rows"]:
        info = selected[row["pn"]]
        net_weight = _quantize(_decimal(info["unit_net_weight"]) * Decimal(row["quantity"]))
        total_net += net_weight
        if info["source_sheet"] != "Sheet1":
            _manual(manual_checks, f"items.{row['pn']}.source_sheet", f"使用 {info['source_sheet']} 数据")
        items.append(
            {
                "pn": row["pn"],
                "quantity": row["quantity"],
                "net_price": _number(row["net_price"]),
                "amount": _number(row["amount"]),
                "description": info["description"],
                "origin_country": info["origin_country"],
                "unit_net_weight": _number(_decimal(info["unit_net_weight"])),
                "net_weight": _number(net_weight),
                "gross_weight": None,
                "source_artifact": info["source_artifact"],
                "source_sheet": info["source_sheet"],
                "source_row": info["source_row"],
            }
        )
    if total_net <= 0:
        return {"status": "needs_input", "missing": ["positive_total_net_weight"]}
    gross_total = _decimal(logistics["gross_weight"])
    assert gross_total is not None
    allocated = Decimal("0")
    for index, item in enumerate(items):
        net = _decimal(item["net_weight"])
        assert net is not None
        gross = gross_total - allocated if index == len(items) - 1 else _quantize(net / total_net * gross_total)
        item["gross_weight"] = _number(gross)
        allocated += gross
    logistics["pieces"] = int(_decimal(logistics["pieces"]))
    logistics["gross_weight"] = _number(gross_total)
    logistics["net_weight"] = _number(total_net)

    canonical = {
        "workflow": WORKFLOW,
        "source_artifacts": {
            "extractions": extraction_artifacts,
            "mineru": source_artifact,
            "order": order_artifact,
            "information": information_artifacts,
            "adjudication": adjudication_artifact,
        },
        "logistics": logistics,
        "items": items,
        "manual_checks": manual_checks,
        "currency": order["currency"],
    }
    _validate_canonical(canonical)
    path = write_json_artifact("tecan_canonical", canonical)
    return {"status": "canonical", "canonical_artifact": path, "manual_checks": manual_checks}


def generate_tecan_documents(canonical_artifact: str) -> dict[str, Any]:
    """Generate the Tecan invoice/packing workbook from one canonical path."""
    canonical = _validate_canonical(read_json_artifact(canonical_artifact))
    workbook = load_workbook(_TEMPLATE, data_only=False)
    customs = workbook["Customs invoice"]
    packing = workbook["Packing List"]
    extra_rows = max(0, len(canonical["items"]) - 25)
    if extra_rows:
        _insert_rows(customs, extra_rows)
        _insert_rows(packing, extra_rows)
    total_row = 66 + extra_rows
    amount_total = Decimal("0")
    customs.cell(24, 8).value = f"( {canonical['currency']})"
    customs.cell(24, 9).value = f"({canonical['currency']})"
    for index, item in enumerate(canonical["items"], start=1):
        row = 32 + index
        for column, value in {
            2: index,
            3: item["pn"],
            4: item["description"],
            6: item["quantity"],
            7: item["origin_country"],
            8: item["net_price"],
            9: item["amount"],
        }.items():
            customs.cell(row, column).value = value
        amount_total += _decimal(item["amount"])
    customs.cell(total_row, 5).value = "Total Amount:"
    customs.cell(total_row, 8).value = canonical["currency"]
    customs.cell(total_row, 9).value = _number(amount_total)

    for index, item in enumerate(canonical["items"], start=1):
        row = 32 + index
        for column, value in {
            2: index,
            3: item["pn"],
            4: item["description"],
            6: item["quantity"],
            7: item["net_weight"],
            8: item["gross_weight"],
        }.items():
            packing.cell(row, column).value = value
    total_pcs_row = 68 + extra_rows
    packing.cell(total_pcs_row, 4).value = "TOTAL PCS:"
    packing.cell(total_pcs_row, 7).value = canonical["logistics"]["pieces"]
    packing.cell(total_pcs_row + 1, 4).value = "TOTAL NET WEIGHT: "
    packing.cell(total_pcs_row + 1, 7).value = canonical["logistics"]["net_weight"]
    packing.cell(total_pcs_row + 2, 4).value = "TOTAL GROSS WEIGHT: "
    packing.cell(total_pcs_row + 2, 7).value = canonical["logistics"]["gross_weight"]
    target = unique_download_path("Tecan_进口_发票箱单_空运", ".xlsx")
    workbook.save(target)
    workbook.close()
    return {
        "status": "generated",
        "canonical_artifact": canonical_artifact,
        "artifacts": [to_virtual_artifact_path(target)],
        "manual_checks": canonical["manual_checks"],
    }


def _validate_extraction(payload: Mapping[str, Any]) -> dict[str, Any]:
    if set(payload) != {"workflow", "extractor", "source_artifact", "logistics", "items"}:
        raise ValueError("Invalid Tecan extraction envelope")
    if payload.get("workflow") != WORKFLOW or payload.get("extractor") not in EXTRACTORS:
        raise ValueError("Invalid Tecan workflow or extractor")
    source = payload.get("source_artifact")
    if not isinstance(source, str) or not source.startswith("/artifacts/"):
        raise ValueError("source_artifact must be an explicit /artifacts/... path")
    logistics = payload.get("logistics")
    if not isinstance(logistics, Mapping) or set(logistics) != set(LOGISTICS_FIELDS):
        raise ValueError("Tecan logistics must contain pieces and gross_weight")
    normalized: dict[str, dict[str, Any]] = {}
    for field_name in LOGISTICS_FIELDS:
        field = logistics[field_name]
        if not isinstance(field, Mapping) or set(field) != {"value", "confidence"}:
            raise ValueError(f"logistics.{field_name} must contain value and confidence")
        confidence = field.get("confidence")
        if confidence not in CONFIDENCES or (field.get("value") is None and confidence != "low"):
            raise ValueError(f"Invalid confidence for logistics.{field_name}")
        normalized[field_name] = {"value": field.get("value"), "confidence": confidence}
    if payload.get("items") != []:
        raise ValueError("Tecan extraction items must be an empty list")
    return {
        "workflow": WORKFLOW,
        "extractor": payload["extractor"],
        "source_artifact": source,
        "logistics": normalized,
        "items": [],
    }


def _load_extractions(paths: Sequence[str]) -> tuple[list[dict[str, Any]], list[str]]:
    payloads: list[dict[str, Any]] = []
    invalid: list[str] = []
    for path in paths:
        try:
            payloads.append(_validate_extraction(read_json_artifact(path)))
        except (OSError, ValueError) as exc:
            invalid.append(f"{path}: {exc}")
    return payloads, invalid


def _load_adjudication(path: str) -> dict[str, Any]:
    payload = read_json_artifact(path)
    if set(payload) != {"workflow", "source_artifacts", "decisions"} or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Tecan adjudication artifact")
    if not isinstance(payload.get("source_artifacts"), list):
        raise ValueError("adjudication source_artifacts must be a list")
    return {**payload, "decisions": _validate_decisions(payload.get("decisions"))}


def _validate_decisions(decisions: Any) -> list[dict[str, Any]]:
    if not isinstance(decisions, list):
        raise ValueError("decisions must be a list")
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for decision in decisions:
        if not isinstance(decision, Mapping) or set(decision) != {"conflict_id", "value", "reason"}:
            raise ValueError("Each decision must contain conflict_id, value and reason")
        conflict_id = decision.get("conflict_id")
        reason = decision.get("reason")
        if not isinstance(conflict_id, str) or not conflict_id or conflict_id in seen:
            raise ValueError("conflict_id must be unique and non-empty")
        if not isinstance(reason, str) or not reason.strip():
            raise ValueError("decision reason must be non-empty")
        seen.add(conflict_id)
        result.append({"conflict_id": conflict_id, "value": decision.get("value"), "reason": reason.strip()})
    return result


def _resolve_vote(
    conflict_id: str,
    field_name: str,
    fields: Sequence[tuple[str, Mapping[str, Any]]],
    decisions: Mapping[str, Any],
    used_decisions: set[str],
    conflicts: list[dict[str, Any]],
    manual_checks: list[dict[str, Any]],
) -> Any:
    values: dict[str, list[tuple[str, Any]]] = {}
    for extractor, field in fields:
        value = field.get("value")
        token = _numeric_token(value)
        if token is not None:
            values.setdefault(token, []).append((extractor, value))
    if not values:
        return None
    winner = max(values.values(), key=len)
    if len(winner) >= 2:
        return _canonical_number(field_name, winner[0][1])
    if len(values) == 1:
        _manual(manual_checks, conflict_id, "仅一个 extractor 提供非空值")
        return _canonical_number(field_name, winner[0][1])
    if conflict_id in decisions:
        used_decisions.add(conflict_id)
        return _canonical_number(field_name, decisions[conflict_id])
    conflicts.append(
        {
            "conflict_id": conflict_id,
            "field": conflict_id,
            "values": [
                {"extractor": extractor, "value": value}
                for entries in values.values()
                for extractor, value in entries
            ],
        }
    )
    return None


def _parse_order_workbook(path: Path, artifact: str) -> dict[str, Any]:
    if path.suffix.lower() != ".xlsx":
        return {"status": "needs_input", "missing": ["order_xlsx"]}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        return {"status": "needs_input", "missing": ["readable_order_xlsx"]}
    try:
        matches = []
        for sheet in workbook.worksheets:
            cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
            headers = {_normalize_header(cell.value): index for index, cell in enumerate(cells)}
            if {"PN", "Order Qty", "Amount"} <= set(headers):
                matches.append((sheet, headers))
        if len(matches) != 1:
            return {"status": "needs_input", "missing": ["unique_order_worksheet"], "artifact": artifact}
        sheet, headers = matches[0]
        rows: list[dict[str, Any]] = []
        currencies: set[str] = set()
        for excel_row, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            pn = normalize_pn(cells[headers["PN"]].value)
            if not pn and all(cells[index].value in (None, "") for index in headers.values()):
                continue
            quantity = _positive_integer(cells[headers["Order Qty"]].value)
            amount_cell = cells[headers["Amount"]]
            amount = _money(amount_cell.value)
            currencies.update(_currencies(amount_cell.value, amount_cell.number_format))
            net_price = None
            if "Net Price" in headers:
                price_cell = cells[headers["Net Price"]]
                net_price = _money(price_cell.value)
                currencies.update(_currencies(price_cell.value, price_cell.number_format))
            if net_price is None and amount is not None and quantity:
                net_price = amount / Decimal(quantity)
            if not pn or quantity is None or amount is None or net_price is None:
                return {"status": "needs_input", "missing": [f"order_row.{excel_row}"]}
            rows.append({"pn": pn, "quantity": quantity, "net_price": net_price, "amount": amount})
        if not rows:
            return {"status": "needs_input", "missing": ["order_rows"]}
        if len(currencies) > 1:
            return {"status": "needs_input", "reason": "order_mixed_currency", "currencies": sorted(currencies)}
        return {"status": "ok", "rows": rows, "currency": next(iter(currencies), "RMB")}
    finally:
        workbook.close()


def _parse_information_workbooks(
    paths: Sequence[tuple[str, Path]],
) -> dict[str, list[dict[str, Any]]] | dict[str, Any]:
    records_by_pn: dict[str, list[dict[str, Any]]] = {}
    for artifact, path in paths:
        if path.suffix.lower() != ".xlsx":
            return {"status": "needs_input", "missing": [f"information_xlsx:{artifact}"]}
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return {"status": "needs_input", "missing": [f"readable_information_xlsx:{artifact}"]}
        try:
            sheets: list[tuple[Any, dict[str, int], str]] = []
            for sheet in workbook.worksheets:
                values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = {_normalize_header(value): index for index, value in enumerate(values)}
                weight = next((name for name in ("净重", "参考净重") if name in headers), None)
                if {"料号", "英文品名", "原产国"} <= set(headers) and weight:
                    sheets.append((sheet, headers, weight))
            if not sheets:
                return {"status": "needs_input", "missing": [f"information_headers:{artifact}"]}
            per_sheet: dict[str, dict[str, list[dict[str, Any]]]] = {}
            for sheet, headers, weight_header in sheets:
                sheet_records: dict[str, list[dict[str, Any]]] = {}
                for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    pn = normalize_pn(values[headers["料号"]] if headers["料号"] < len(values) else None)
                    if not pn:
                        continue
                    description = _text(values[headers["英文品名"]] if headers["英文品名"] < len(values) else None)
                    country = _text(values[headers["原产国"]] if headers["原产国"] < len(values) else None)
                    weight = _decimal(values[headers[weight_header]] if headers[weight_header] < len(values) else None)
                    if not description or not country or weight is None or weight <= 0:
                        continue
                    sheet_records.setdefault(pn, []).append(
                        {
                            "description": description,
                            "origin_country": country,
                            "unit_net_weight": _number(weight),
                            "source_artifact": artifact,
                            "source_sheet": sheet.title,
                            "source_row": row_number,
                        }
                    )
                per_sheet[sheet.title] = sheet_records
            all_pns = {pn for records in per_sheet.values() for pn in records}
            for pn in all_pns:
                primary = per_sheet.get("Sheet1", {}).get(pn, [])
                candidates = primary or [item for name, records in per_sheet.items() if name != "Sheet1" for item in records.get(pn, [])]
                records_by_pn.setdefault(pn, []).extend(candidates)
        finally:
            workbook.close()
    return records_by_pn


def _distinct_info_records(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    keys: set[tuple[str, str, str]] = set()
    for record in records:
        key = (
            record["description"].casefold(),
            record["origin_country"].casefold(),
            _numeric_token(record["unit_net_weight"]) or "",
        )
        if key not in keys:
            keys.add(key)
            result.append(record)
    return result


def _select_info_record(records: Sequence[dict[str, Any]], choice: str | None) -> dict[str, Any] | None:
    text = str(choice or "").strip().casefold().replace(" ", "")
    if not text:
        return None
    matches = [record for record in records if text == record["source_artifact"].casefold().replace(" ", "")]
    if not matches:
        aliases = (
            ("配件", "part")
            if "配件" in text or "part" in text
            else ("设备", "device", "equipment")
            if any(token in text for token in ("设备", "device", "equipment"))
            else (text,)
        )
        matches = [
            record
            for record in records
            if any(alias in Path(record["source_artifact"]).name.casefold() for alias in aliases)
        ]
    return matches[0] if len(matches) == 1 else None


def _validate_canonical(payload: Mapping[str, Any]) -> dict[str, Any]:
    if set(payload) != CANONICAL_KEYS or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Tecan canonical contract")
    source = payload.get("source_artifacts")
    if not isinstance(source, Mapping) or set(source) != {"extractions", "mineru", "order", "information", "adjudication"}:
        raise ValueError("Invalid Tecan canonical source_artifacts")
    logistics = payload.get("logistics")
    if not isinstance(logistics, Mapping) or set(logistics) != {"pieces", "gross_weight", "net_weight"}:
        raise ValueError("Invalid Tecan canonical logistics")
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise ValueError("Tecan canonical items must not be empty")
    if any(not isinstance(item, Mapping) or set(item) != ITEM_KEYS for item in items):
        raise ValueError("Invalid Tecan canonical item")
    if not isinstance(payload.get("manual_checks"), list) or not isinstance(payload.get("currency"), str):
        raise ValueError("Invalid Tecan canonical metadata")
    return dict(payload)


def _insert_rows(sheet: Any, count: int) -> None:
    sheet.insert_rows(58, amount=count)
    for row in range(58, 58 + count):
        source_row = 57
        sheet.row_dimensions[row].height = sheet.row_dimensions[source_row].height
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(row, column)
            if source.has_style:
                target.font = copy.copy(source.font)
                target.fill = copy.copy(source.fill)
                target.border = copy.copy(source.border)
                target.alignment = copy.copy(source.alignment)
                target.protection = copy.copy(source.protection)
            target.number_format = source.number_format


def normalize_pn(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        if not value.is_finite():
            return ""
        value = value.quantize(Decimal("1")) if value == value.to_integral_value() else value
        return format(value, "f")
    text = re.sub(r"\s+", "", str(value).strip())
    match = re.fullmatch(r"([+-]?\d+)\.0+", text)
    return match.group(1) if match else text


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _positive_integer(value: Any) -> int | None:
    number = _decimal(value)
    if number is None or number <= 0 or number != number.to_integral_value():
        return None
    return int(number)


def _money(value: Any) -> Decimal | None:
    number = _decimal(value)
    if number is not None:
        return number
    if not isinstance(value, str):
        return None
    candidate = "".join(value.split()).replace("，", ",").replace("．", ".")
    candidate = _CURRENCY_RE.sub("", candidate).replace(",", "")
    try:
        result = Decimal(candidate)
        return result if result.is_finite() else None
    except InvalidOperation:
        return None


def _currencies(value: Any, number_format: str | None) -> set[str]:
    result: set[str] = set()
    text = f"{value if isinstance(value, str) else ''} {number_format or ''}".upper()
    for token in _CURRENCY_RE.findall(text):
        normalized = token.upper()
        if normalized in {"¥", "￥", "RMB", "CNY"}:
            result.add("RMB")
        elif normalized in {"$", "US$", "USD"}:
            result.add("USD")
        elif normalized in {"€", "EUR"}:
            result.add("EUR")
        elif normalized in {"£", "GBP"}:
            result.add("GBP")
        else:
            result.add(normalized)
    return result


def _canonical_number(field: str, value: Any) -> Any:
    number = _decimal(value)
    if number is None or number <= 0:
        return None
    if field == "pieces":
        return int(number) if number == number.to_integral_value() else None
    return _number(number)


def _numeric_token(value: Any) -> str | None:
    number = _decimal(value)
    return format(number.normalize(), "f") if number is not None and number > 0 else None


def _decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return Decimal(str(value))
    text = str(value).strip()
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
    if not match:
        return None
    try:
        result = Decimal(match.group(0).replace(",", ""))
        return result if result.is_finite() else None
    except InvalidOperation:
        return None


def _number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _quantize(value: Decimal) -> Decimal:
    return value.quantize(_WEIGHT_QUANT, rounding=ROUND_HALF_UP)


def _text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _needs_c(source_artifact: str, reasons: Sequence[str]) -> dict[str, Any]:
    return {"status": "needs_c", "source_artifact": source_artifact, "reasons": list(reasons)}


def _manual(checks: list[dict[str, Any]], field: str, reason: str) -> None:
    if not any(item.get("field") == field and item.get("reason") == reason for item in checks):
        checks.append({"field": field, "reason": reason})
