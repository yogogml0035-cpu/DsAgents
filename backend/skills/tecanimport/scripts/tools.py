"""Tecan 境外业务的 XLSX 读取与终态 JSON 校验工具。"""
from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from integrations.artifacts import resolve_artifact_path, write_json_artifact
from skills.tecanimport.schema import TecanOverseasRecognitionResult


FINALIZE_TECAN_RESULT_TOOL = "finalize_tecan_overseas_recognition"


def inspect_supply_chain_workbooks(file_paths: list[str]) -> dict[str, Any]:
    """将显式 XLSX 原样转为可读取 JSON artifact，供渠道 Skill 动态识别材料角色。"""
    if not file_paths:
        raise ValueError("file_paths must not be empty")

    workbooks: list[dict[str, str]] = []
    problems: list[dict[str, str]] = []
    for raw_path in file_paths:
        try:
            source = resolve_artifact_path(raw_path)
            if source.suffix.lower() != ".xlsx" or not source.is_file():
                raise ValueError("只支持存在的 .xlsx 文件")
            payload = _workbook_payload(source, raw_path)
            workbooks.append(
                {
                    "source_artifact": raw_path,
                    "result_path": write_json_artifact("tecan_workbook", payload),
                }
            )
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            problems.append(
                {
                    "source": raw_path,
                    "location": "workbook",
                    "issue": str(exc),
                    "action": "提供可读取的 .xlsx，或忽略该不支持材料",
                }
            )
    return {"workbooks": workbooks, "problems": problems}


def finalize_tecan_overseas_recognition(result: TecanOverseasRecognitionResult) -> str:
    """校验并返回 Tecan 最终业务 JSON；不写 Excel 或候选 artifact。"""
    validated = TecanOverseasRecognitionResult.model_validate(result)
    return json.dumps(validated.model_dump(mode="json"), ensure_ascii=False, separators=(",", ":"))


def _workbook_payload(path: Path, artifact: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValueError("XLSX 无法读取") from exc
    try:
        return {
            "source_artifact": artifact,
            "sheets": [
                {
                    "name": sheet.title,
                    "rows": [
                        row
                        for values in sheet.iter_rows(values_only=True)
                        if (row := _row_values(values)) is not None
                    ],
                }
                for sheet in workbook.worksheets
            ],
        }
    finally:
        workbook.close()


def _row_values(values: tuple[Any, ...]) -> list[Any] | None:
    row = [_cell_value(value) for value in values]
    while row and row[-1] is None:
        row.pop()
    return row if any(value is not None for value in row) else None


def _cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    return value
