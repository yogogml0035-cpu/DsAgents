from __future__ import annotations

import os
import re
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from integrations.artifacts import resolve_artifact_path


VALID_TRACKING_STATUSES = {"进口", "出1", "出2", "已出"}
# English keys match PhilipsWgqRecognitionResult item fields (tool + API contract).
MASTER_FIELDS = (
    "chinese_name",
    "specification",
    "origin_country",
    "customs_code",
    "declaration_elements",
    "unit",
    "legal_unit_1",
    "legal_unit_2",
    "new_or_used",
    "business_unit",
)
ORACLE_FIELDS = (
    "chinese_name",
    "specification",
    "origin_country",
    "customs_code",
    "unit",
    "legal_unit_1",
    "legal_unit_2",
)

# Keys are English master-data fields; values are Tracking Excel header aliases.
_ALIASES = {
    "product_id": ("飞利浦料号", "料号", "12NC"),
    "chinese_name": ("中文品名",),
    "specification": ("规格型号", "型号"),
    "origin_country": ("原产国",),
    "customs_code": ("HS code", "HS编码", "海关编码"),
    "declaration_elements": ("申报要素",),
    "unit": ("申报计量单位", "单位"),
    "legal_unit_1": ("法定第一单位", "法定单位"),
    "legal_unit_2": ("法定第二单位",),
    "new_or_used": ("新旧",),
    "business_unit": ("Modality", "BU"),
}
_DECLARATION_ELEMENT_HEADERS = {
    "用途",
    "原理",
    "品牌",
    "型号",
    "主要配置",
    "新旧",
    "探头数量",
    "何种类型超声波（仅针对彩超机及其配件）",
    "热容量（仅针对X射线管）",
    "容量（针对电池）",
    "是否含汞（针对电池）",
    "额定电压",
    "结构类型（是否为同轴电缆）",
    "单个长度",
}

_ORACLE_SQL = """
select
    c.zwmc,
    c.ggxh,
    c.ycg,
    c.hsbm,
    c.jldw,
    u1.unit_name,
    u2.unit_name
from od.chda c
left join dongsong.good g on g.goodcode = '01' || c.hsbm
left join dongsong.custom_unit u1 on u1.unit_code = g.unit
left join dongsong.custom_unit u2 on u2.unit_code = g.unit2
where c.chbm = :product_id
"""
_ORACLE_CLIENT_INITIALIZED = False


def lookup_philips_wgq_master_data(
    product_ids: list[str],
    tracking_artifact: str | None = None,
) -> dict[str, Any]:
    """按唯一 12NC 查询共享主数据；WGQ 可额外传 Tracking，Oracle 只补稳定字段。"""
    normalized_ids = list(dict.fromkeys(filter(None, (normalize_product_id(value) for value in product_ids))))
    items = {product_id: {field: None for field in MASTER_FIELDS} for product_id in normalized_ids}
    problems: list[dict[str, str]] = []

    if tracking_artifact and normalized_ids:
        try:
            tracking = _tracking_data(resolve_artifact_path(tracking_artifact), set(normalized_ids))
        except Exception as exc:
            problems.append(_problem(tracking_artifact, "tracking", f"Tracking 读取失败：{_error_text(exc)}", "检查 Tracking .xlsx 后重试"))
        else:
            for product_id, values in tracking.items():
                items[product_id].update(values)
            for product_id in normalized_ids:
                if product_id not in tracking:
                    problems.append(_problem(tracking_artifact, f"items.{product_id}", "未找到合格 Tracking 进口行", "核对 12NC 或补充 Tracking"))

    needs_oracle = [
        product_id
        for product_id, values in items.items()
        if any(values[field] is None for field in ORACLE_FIELDS)
    ]
    oracle, oracle_problems = _oracle_data(needs_oracle)
    problems.extend(oracle_problems)
    for product_id, values in oracle.items():
        for field, value in values.items():
            if items[product_id][field] is None and value is not None:
                items[product_id][field] = value

    return {
        "items": [{"product_id": product_id, **items[product_id]} for product_id in normalized_ids],
        "problems": problems,
    }


def normalize_product_id(value: Any) -> str:
    text = _cell_text(value) or ""
    cleaned = re.sub(r"[\s-]+", "", text)
    return cleaned[-12:] if len(cleaned) > 12 else cleaned


def _tracking_data(path: Path, requested_ids: set[str]) -> dict[str, dict[str, str | None]]:
    if path.suffix.lower() != ".xlsx" or not path.is_file():
        raise ValueError("Tracking 必须是存在的 .xlsx 文件")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "进口" not in workbook.sheetnames:
            raise ValueError("Tracking 缺少 进口 sheet")
        import_header, import_rows = _sheet_rows(workbook["进口"])
        import_indexes = _header_indexes(import_header)
        product_index = _find_index(import_indexes, _ALIASES["product_id"])
        if product_index is None:
            raise ValueError("进口 sheet 缺少料号列")

        selected: dict[str, Sequence[Any]] = {}
        for row in reversed(import_rows):
            status = _cell_text(row[0] if row else None)
            product_id = normalize_product_id(_at(row, product_index))
            if status in VALID_TRACKING_STATUSES and product_id in requested_ids and product_id not in selected:
                selected[product_id] = row

        declaration_rows: dict[str, Sequence[Any]] = {}
        declaration_header: Sequence[Any] = ()
        declaration_indexes: dict[str, int] = {}
        if selected and "申报要素" in workbook.sheetnames:
            declaration_header, rows = _sheet_rows(workbook["申报要素"])
            declaration_indexes = _header_indexes(declaration_header)
            declaration_product_index = _find_index(declaration_indexes, _ALIASES["product_id"])
            if declaration_product_index is not None:
                for row in rows:
                    product_id = normalize_product_id(_at(row, declaration_product_index))
                    if product_id in selected and product_id not in declaration_rows:
                        declaration_rows[product_id] = row

        result: dict[str, dict[str, str | None]] = {}
        for product_id, import_row in selected.items():
            declaration_row = declaration_rows.get(product_id)
            values: dict[str, str | None] = {}
            for field in MASTER_FIELDS:
                declaration_value = (
                    _field_value(field, declaration_header, declaration_indexes, declaration_row)
                    if declaration_row is not None
                    else None
                )
                values[field] = declaration_value or _field_value(
                    field, import_header, import_indexes, import_row
                )
            result[product_id] = values
        return result
    finally:
        workbook.close()


def _sheet_rows(sheet: Any) -> tuple[tuple[Any, ...], list[tuple[Any, ...]]]:
    header = tuple(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    width = max((index for index, value in enumerate(header) if value is not None), default=-1) + 1
    if width == 0:
        raise ValueError(f"{sheet.title} sheet 表头为空")
    return header[:width], list(sheet.iter_rows(min_row=2, max_col=width, values_only=True))


def _header_indexes(header: Sequence[Any]) -> dict[str, int]:
    return {
        _header_key(value): index
        for index, value in enumerate(header)
        if _cell_text(value) is not None
    }


def _find_index(indexes: Mapping[str, int], aliases: Sequence[str]) -> int | None:
    return next((indexes[_header_key(alias)] for alias in aliases if _header_key(alias) in indexes), None)


def _field_value(
    field: str,
    header: Sequence[Any],
    indexes: Mapping[str, int],
    row: Sequence[Any],
) -> str | None:
    for alias in _ALIASES[field]:
        index = indexes.get(_header_key(alias))
        value = _cell_text(_at(row, index)) if index is not None else None
        if value is not None:
            return value
    if field != "declaration_elements":
        return None
    elements = []
    for index, name in enumerate(header):
        label = _cell_text(name)
        value = _cell_text(_at(row, index))
        if label in _DECLARATION_ELEMENT_HEADERS and value is not None:
            elements.append(f"{label}：{value}")
    return "；".join(elements) or None


def _oracle_data(product_ids: Sequence[str]) -> tuple[dict[str, dict[str, str | None]], list[dict[str, str]]]:
    if not product_ids:
        return {}, []
    dsn, username, password = (os.getenv(name) for name in ("ORACLE_DSN", "ORACLE_USERNAME", "ORACLE_PASSWORD"))
    if not all((dsn, username, password)):
        return {}, [_problem("Oracle", "master_data", "Oracle 配置缺失", "配置 Oracle 后重试，或人工补齐主数据")]
    try:
        import oracledb

        _init_oracle_client(os.getenv("ORACLE_CLIENT_LIB_DIR"))
        timeout = float(os.getenv("ORACLE_TIMEOUT_SECONDS") or "30")
        result: dict[str, dict[str, str | None]] = {}
        problems: list[dict[str, str]] = []
        with oracledb.connect(user=username, password=password, dsn=dsn, tcp_connect_timeout=timeout) as connection:
            connection.call_timeout = int(timeout * 1000)
            with connection.cursor() as cursor:
                for product_id in product_ids:
                    cursor.execute(_ORACLE_SQL, product_id=product_id)
                    row = cursor.fetchone()
                    if row is None:
                        problems.append(_problem("Oracle", f"items.{product_id}", "Oracle 未命中 12NC", "人工补齐主数据"))
                        continue
                    result[product_id] = {
                        field: _cell_text(_at(row, index))
                        for index, field in enumerate(ORACLE_FIELDS)
                    }
        return result, problems
    except Exception as exc:
        return {}, [_problem("Oracle", "master_data", f"Oracle 查询失败：{_error_text(exc)}", "检查 Oracle 后重试，或人工补齐主数据")]


def _init_oracle_client(lib_dir: str | None) -> None:
    global _ORACLE_CLIENT_INITIALIZED
    if _ORACLE_CLIENT_INITIALIZED:
        return
    lib_dir = lib_dir or _bundled_oracle_client_dir()
    if not lib_dir:
        return
    import oracledb

    oracledb.init_oracle_client(lib_dir=lib_dir)
    _ORACLE_CLIENT_INITIALIZED = True


def _bundled_oracle_client_dir() -> str | None:
    """Use the checked-in Windows Instant Client when no explicit path is configured."""
    if os.name != "nt":
        return None
    client_dir = Path(__file__).resolve().parents[3] / ".oracle" / "instantclient" / "instantclient_19_31"
    return str(client_dir) if (client_dir / "oci.dll").is_file() else None


def _problem(source: str, location: str, issue: str, action: str) -> dict[str, str]:
    return {"source": source, "location": location, "issue": issue, "action": action}


def _header_key(value: Any) -> str:
    return re.sub(r"\s+", "", _cell_text(value) or "").casefold()


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.upper() in {"//", "/", "-", "--", "N/A", "NA"} or text.startswith(("未找到", "需确认")):
        return None
    return text or None


def _at(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def _error_text(exc: Exception) -> str:
    return str(exc).strip() or exc.__class__.__name__
