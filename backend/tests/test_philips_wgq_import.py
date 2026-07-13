from __future__ import annotations

import copy
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from integrations.artifacts import read_json_artifact, resolve_artifact_path
from skills.philipswgqimport.scripts.tools import (
    _validate_extraction,
    generate_philips_wgq_import,
    save_philips_wgq_extraction,
)


def run() -> None:
    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
        artifacts = (Path(tmp) / "artifacts").resolve()
        source = artifacts / "downloads" / "mineru.json"
        source.parent.mkdir(parents=True)
        source.write_text('{"content_list":[]}', encoding="utf-8")
        tracking = artifacts / "uploads" / "tracking.xlsx"
        tracking.parent.mkdir(parents=True)
        _tracking_fixture(tracking)

        with patch("integrations.artifacts.artifacts_root", return_value=artifacts):
            base = _extraction()
            a = _save(base, "philips-wgq-extractor-a")
            b = _save(base, "philips-wgq-extractor-b")

            # A/B consistent -> one-shot generate produces 3 Excel.
            consistent = _generate([a, b])
            assert consistent["status"] == "generated"
            canonical = read_json_artifact(consistent["canonical_artifact"])
            item = canonical["items"][0]
            assert item["product_id"] == "10989000085103"
            assert item["normalized_product_id"] == "989000085103"
            assert item["quantity"] == 2
            assert item["total_price"] == 20
            assert item["gross_weight"] == 13.5
            assert item["net_weight"] == 11
            assert item["po_number"] == "PO1/PO2"
            assert canonical["logistics"]["gross_weight"] == 13.5
            assert canonical["international_forwarder"] == "DHL"

            # Only one A/B extractor -> needs C, surfaced as input_problems.
            one_failed = _generate([a, "/artifacts/downloads/missing.json"])
            assert one_failed["code"] == "input_problems"
            assert any("extractor C" in p["issue"] for p in one_failed["problems"])

            # A/B conflict on a field -> needs C, surfaced as input_problems.
            conflict_payload = copy.deepcopy(base)
            conflict_payload["items"][0]["description"] = _field("different")
            conflict_b = _save(conflict_payload, "philips-wgq-extractor-b")
            conflict_result = _generate([a, conflict_b])
            assert conflict_result["code"] == "input_problems"
            assert any("extractor C" in p["issue"] for p in conflict_result["problems"])

            # Premature decisions on A/B conflict -> input_problems (C required first).
            premature = _generate(
                [a, conflict_b],
                decisions=[{"conflict_id": "items.row-1.description", "value": "Medical part", "reason": "回查"}],
            )
            assert premature["code"] == "input_problems"
            assert any("C 抽取尚未完成" in p["issue"] for p in premature["problems"])

            # C alone (consistent with the original source) -> generate.
            c = _save(base, "philips-wgq-extractor-c")
            assert _generate([c])["status"] == "generated"

            # A/B/C majority on a logistics field -> generate.
            majority_b = copy.deepcopy(base)
            majority_b["logistics"]["hawb_number"] = _field("B-H")
            majority_b_path = _save(majority_b, "philips-wgq-extractor-b")
            majority_c = copy.deepcopy(base)
            majority_c["logistics"]["hawb_number"] = _field("J180317")
            majority_c_path = _save(majority_c, "philips-wgq-extractor-c")
            assert _generate([a, majority_b_path, majority_c_path])["status"] == "generated"

            # A/B/C conflict with no decision -> input_problems listing the conflict.
            no_majority_c = copy.deepcopy(base)
            no_majority_c["logistics"]["hawb_number"] = _field("C-H")
            no_majority_c_path = _save(no_majority_c, "philips-wgq-extractor-c")
            unresolved = _generate([a, majority_b_path, no_majority_c_path])
            assert unresolved["code"] == "input_problems"
            assert any(
                p["location"] == "logistics.hawb_number" and "不一致" in p["issue"]
                for p in unresolved["problems"]
            )

            # Same A/B/C conflict resolved by an inline decision -> generate.
            resolved = _generate(
                [a, majority_b_path, no_majority_c_path],
                decisions=[
                    {"conflict_id": "logistics.hawb_number", "value": "J180317", "reason": "回查确认"}
                ],
            )
            assert resolved["status"] == "generated"

            # Decisions reference a non-existent conflict_id -> input_problems.
            bad_decision = _generate(
                [a, majority_b_path, majority_c_path],
                decisions=[{"conflict_id": "items.row-1.description", "value": "x", "reason": "nope"}],
            )
            assert bad_decision["code"] == "input_problems"
            assert any(p["location"] == "decisions" for p in bad_decision["problems"])

            # All extractors jointly missing a logistics field -> input_problems.
            missing = copy.deepcopy(base)
            missing["logistics"]["gross_weight"] = _field(None, "low")
            missing_a = _save(missing, "philips-wgq-extractor-a")
            missing_b = _save(missing, "philips-wgq-extractor-b")
            missing_c = _save(missing, "philips-wgq-extractor-c")
            required = _generate([missing_a, missing_b, missing_c])
            assert required["code"] == "input_problems"
            assert any(p["location"] == "logistics.gross_weight" for p in required["problems"])

            # No valid items -> input_problems.
            empty = {"logistics": base["logistics"], "items": []}
            empty_a = _save(empty, "philips-wgq-extractor-a")
            empty_b = _save(empty, "philips-wgq-extractor-b")
            empty_c = _save(empty, "philips-wgq-extractor-c")
            empty_result = _generate([empty_a, empty_b, empty_c])
            assert empty_result["code"] == "input_problems"
            assert any("商品行" in p["issue"] for p in empty_result["problems"])

            # Missing forwarder -> input_problems.
            no_forwarder = generate_philips_wgq_import(
                extraction_artifacts=[a, b],
                tracking_artifact="/artifacts/uploads/tracking.xlsx",
                international_forwarder=None,
            )
            assert no_forwarder["code"] == "input_problems"
            assert any(p["location"] == "international_forwarder" for p in no_forwarder["problems"])

            # Old extraction contract is rejected.
            try:
                _validate_extraction(
                    {
                        "schema_version": 1,
                        "extractor_id": "philips-wgq-extractor-a",
                        "logistics_fields": {},
                        "raw_items": [],
                    }
                )
            except ValueError:
                pass
            else:
                raise AssertionError("old Philips extraction contract must be rejected")

            # Generated workbooks carry the expected fields (Oracle config absent
            # => manual check for oracle_units is present).
            with patch.dict(os.environ, {}, clear=True):
                _check_workbooks(consistent["artifacts"])
            assert any(check["field"] == "oracle_units" for check in consistent["manual_checks"])


def _field(value: object, confidence: str = "high") -> dict[str, object]:
    return {"value": value, "confidence": confidence}


def _extraction() -> dict[str, object]:
    logistics = {
        "hawb_number": _field("J180317"),
        "pieces": _field(1),
        "gross_weight": _field("13.5 KG"),
        "shipper_country": _field("DE"),
    }
    first = {
        "product_id_raw": _field("109890-000-85103"),
        "description": _field("Medical part"),
        "quantity": _field(1),
        "unit_price": _field("10.00"),
        "total_price": _field("10.00"),
        "currency": _field("US Dollar"),
        "po_number": _field("PO 1"),
        "raw_country": _field("USA"),
        "gross_weight": _field(6),
    }
    second = copy.deepcopy(first)
    second["po_number"] = _field("PO 2")
    second["gross_weight"] = _field(7.5)
    return {"logistics": logistics, "items": [first, second]}


def _save(payload: dict[str, object], extractor: str) -> str:
    return save_philips_wgq_extraction(
        extractor,
        "/artifacts/downloads/mineru.json",
        payload["logistics"],
        payload["items"],
    )["artifact_path"]


def _generate(paths: list[str], **kwargs: object) -> dict[str, object]:
    return generate_philips_wgq_import(
        extraction_artifacts=paths,
        tracking_artifact="/artifacts/uploads/tracking.xlsx",
        international_forwarder="DHL",
        customs_mode="普货",
        **kwargs,
    )


def _tracking_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "进口"
    headers = [
        "状态", "料号", "数量", "型号", "币种", "备案单价USD", "备案总价USD", "原产国",
        "运单号", "国际货代", "进境STO(PO)", "件数", "净重", "毛重",
        "Figo提供完整申报要素日期", "HS编码", "中文品名", "法定第一单位", "Modality",
        "监管条件", "监管条件详解", "海关税则书所列商品名称", "进境备注",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "已出", "989000085103", 2, "old", "USD", 10, 20, "美国",
            "OLD", "DHL", "OLDPO", 2, 8, 10, None,
            "901890", "医疗部件", "个", "CT", "A", "detail", "medical device", "进口",
        ]
    )
    declaration = workbook.create_sheet("申报要素")
    declaration.append(["飞利浦料号", "HS code", "中文品名", "法定第一单位"])
    declaration.append(["989000085103", "901890", "医疗部件", "个"])
    workbook.save(path)


def _check_workbooks(paths: list[str]) -> None:
    resolved = [resolve_artifact_path(path) for path in paths]
    tracking = next(path for path in resolved if "进境更新" in path.name)
    invoice = next(path for path in resolved if "invoice_packing" in path.name)
    bonded = next(path for path in resolved if "核注清单" in path.name)

    workbook = load_workbook(tracking, data_only=False)
    assert workbook["进口"].max_row == 3
    assert workbook["进口"]["B3"].value == "989000085103"
    assert isinstance(workbook["进口"]["N3"].value, (int, float))
    assert workbook["进口"]["P3"].value == "901890"
    assert workbook["进口"]["S3"].value == "CT"
    workbook.close()

    workbook = load_workbook(invoice, data_only=False)
    assert set(workbook.sheetnames) >= {"Customs invoice", "Packing List"}
    assert workbook["Customs invoice"]["K32"].value == 20
    assert workbook["Packing List"]["J37"].value == 13.5
    assert workbook["Packing List"]["F45"].value == 11
    workbook.close()

    workbook = load_workbook(bonded, data_only=False)
    assert set(workbook.sheetnames) >= {"表头", "表体"}
    assert workbook["表体"]["F2"].value == "989000085103"
    assert isinstance(workbook["表体"]["T2"].value, (int, float))
    assert workbook["表体"]["J2"].value == "需确认：申报计量单位"
    workbook.close()


if __name__ == "__main__":
    run()
