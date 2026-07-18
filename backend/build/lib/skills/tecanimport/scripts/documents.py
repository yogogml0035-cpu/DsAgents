"""Tecan 帝肯进口确定性 Excel 写入：发票箱单（Customs invoice + Packing List）。

只消费已经过校验的 canonical dict；不做业务判断、不做抽取。
"""
from __future__ import annotations

import copy
from decimal import Decimal
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from integrations.artifacts import to_virtual_artifact_path, unique_download_path


_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "tecan-import"
    / "assets"
    / "Tecan_进口_发票箱单_空运.xlsx"
)


def generate_invoice_packing(canonical: Mapping[str, Any]) -> str:
    """根据一份 canonical 填写帝肯模板的 Customs invoice 与 Packing List。"""
    workbook = load_workbook(_TEMPLATE, data_only=False)
    try:
        customs = workbook["Customs invoice"]
        packing = workbook["Packing List"]
        extra_rows = max(0, len(canonical["items"]) - 25)
        if extra_rows:
            insert_rows(customs, extra_rows)
            insert_rows(packing, extra_rows)
        total_row = 66 + extra_rows
        amount_total = Decimal("0")
        customs.cell(24, 8).value = f"( {canonical['currency']})"
        customs.cell(24, 9).value = f"({canonical['currency']})"
        for index, item in enumerate(canonical["items"], start=1):
            row = 32 + index
            for column, value in {
                2: index,
                3: item["pn"],
                4: item["description"],
                6: item["quantity"],
                7: item["origin_country"],
                8: item["net_price"],
                9: item["amount"],
            }.items():
                customs.cell(row, column).value = value
            amount_total += to_decimal(item["amount"])
        customs.cell(total_row, 5).value = "Total Amount:"
        customs.cell(total_row, 8).value = canonical["currency"]
        customs.cell(total_row, 9).value = to_number(amount_total)

        for index, item in enumerate(canonical["items"], start=1):
            row = 32 + index
            for column, value in {
                2: index,
                3: item["pn"],
                4: item["description"],
                6: item["quantity"],
                7: item["net_weight"],
                8: item["gross_weight"],
            }.items():
                packing.cell(row, column).value = value
        total_pcs_row = 68 + extra_rows
        packing.cell(total_pcs_row, 4).value = "TOTAL PCS:"
        packing.cell(total_pcs_row, 7).value = canonical["logistics"]["pieces"]
        packing.cell(total_pcs_row + 1, 4).value = "TOTAL NET WEIGHT: "
        packing.cell(total_pcs_row + 1, 7).value = canonical["logistics"]["net_weight"]
        packing.cell(total_pcs_row + 2, 4).value = "TOTAL GROSS WEIGHT: "
        packing.cell(total_pcs_row + 2, 7).value = canonical["logistics"]["gross_weight"]
        target = unique_download_path("Tecan_进口_发票箱单_空运", ".xlsx")
        workbook.save(target)
    finally:
        workbook.close()
    return to_virtual_artifact_path(target)


def insert_rows(sheet: Any, count: int) -> None:
    sheet.insert_rows(58, amount=count)
    source_row = 57
    for row in range(58, 58 + count):
        sheet.row_dimensions[row].height = sheet.row_dimensions[source_row].height
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(row, column)
            if source.has_style:
                target.font = copy.copy(source.font)
                target.fill = copy.copy(source.fill)
                target.border = copy.copy(source.border)
                target.alignment = copy.copy(source.alignment)
                target.protection = copy.copy(source.protection)
            target.number_format = source.number_format


def to_decimal(value: Any) -> Decimal:
    from skills.tecanimport.scripts.tools import _decimal

    return _decimal(value) or Decimal("0")


def to_number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)
