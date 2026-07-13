"""Tecan 帝肯进口业务：抽取保存 + 一站式 canonical/订单信息表 join/计算/生成。

只暴露两个 Tool：
- save_tecan_extraction：规范化并保存一次 A/B/C 物流抽取结果（items 必须为空）。
- generate_tecan_import：接收抽取结果与裁决 decisions，一次完成物流校验、订单/
  信息表 join、净/毛重计算和发票箱单写入。

业务问题统一以 {"code": "input_problems", "problems": [...]} 返回，run 结束。
"""
from __future__ import annotations

import math
import re
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping, Sequence
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from dsagents.integrations.artifacts import (
    read_json_artifact,
    resolve_artifact_path,
    write_json_artifact,
)
from dsagents.skills.tecanimport.scripts import documents as docs


WORKFLOW = "tecan-import"
EXTRACTORS = frozenset(f"tecan-extractor-{name}" for name in "abc")
CONFIDENCES = frozenset({"high", "medium", "low"})
LOGISTICS_FIELDS = ("pieces", "gross_weight")
CANONICAL_KEYS = frozenset(
    {"workflow", "source_artifacts", "logistics", "items", "manual_checks", "currency"}
)
ITEM_KEYS = frozenset(
    {
        "pn",
        "quantity",
        "net_price",
        "amount",
        "description",
        "origin_country",
        "unit_net_weight",
        "net_weight",
        "gross_weight",
        "source_artifact",
        "source_sheet",
        "source_row",
    }
)
_WEIGHT_QUANT = Decimal("0.001")
_CURRENCY_RE = re.compile(r"US\$|RMB|CNY|USD|EUR|GBP|JPY|SGD|HKD|CHF|AUD|CAD|[¥￥$€£]", re.IGNORECASE)


def save_tecan_extraction(
    extractor: str,
    source_artifact: str,
    logistics: dict[str, dict[str, Any]],
    items: list[dict[str, Any]],
) -> dict[str, str]:
    """Validate and immutably save one Tecan logistics extraction artifact."""
    if not resolve_artifact_path(source_artifact).is_file():
        raise ValueError(f"source_artifact not found: {source_artifact}")
    payload = _validate_extraction(
        {
            "workflow": WORKFLOW,
            "extractor": extractor,
            "source_artifact": source_artifact,
            "logistics": logistics,
            "items": items,
        }
    )
    path = write_json_artifact(extractor, payload)
    return {"extractor": extractor, "artifact_path": path}


def generate_tecan_import(
    extraction_artifacts: list[str],
    order_artifact: str,
    information_artifacts: list[str],
    decisions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """One-shot: reconcile logistics + decisions → canonical → 发票箱单.

    Returns one of:
      {"status": "generated", "canonical_artifact", "artifacts", "manual_checks"}
      {"code": "input_problems", "problems": [{source, location, issue, action}]}
    """
    built = _build_canonical(
        extraction_artifacts=extraction_artifacts,
        order_artifact=order_artifact,
        information_artifacts=information_artifacts,
        decisions=_validate_decisions(decisions or []),
    )
    if built.get("code") == "input_problems":
        return built
    canonical = built["canonical"]
    artifact = docs.generate_invoice_packing(canonical)
    return {
        "status": "generated",
        "canonical_artifact": built["canonical_artifact"],
        "artifacts": [artifact],
        "manual_checks": canonical["manual_checks"],
    }


# ---------- canonical 构建 ----------


def _build_canonical(
    *,
    extraction_artifacts: list[str],
    order_artifact: str,
    information_artifacts: list[str],
    decisions: list[dict[str, Any]],
) -> dict[str, Any]:
    problems: list[dict[str, Any]] = []

    try:
        order_path = resolve_artifact_path(order_artifact)
    except ValueError as exc:
        problems.append(_problem(order_artifact, "order_artifact", str(exc), "提供正确的订单 artifact 路径"))
        order_path = None
    if order_path is not None and not order_path.is_file():
        problems.append(_problem(order_artifact, "order_artifact", "订单文件不存在", "提供正确的订单 Excel"))
        order_path = None

    info_paths: list[tuple[str, Path]] = []
    for raw in information_artifacts:
        try:
            info_paths.append((raw, resolve_artifact_path(raw)))
        except ValueError as exc:
            problems.append(_problem(raw, "information_artifacts", str(exc), "提供正确的信息表路径"))
    if not info_paths:
        problems.append(
            _problem(information_artifacts, "information_artifacts", "缺少信息表", "至少提供一个信息表 artifact")
        )

    payloads, invalid = _load_extractions(extraction_artifacts)
    for reason in invalid:
        problems.append(_problem(extraction_artifacts, "extraction", reason, "重新保存合法抽取结果"))

    if payloads:
        extractor_ids = [payload["extractor"] for payload in payloads]
        if len(set(extractor_ids)) != len(extractor_ids):
            problems.append(
                _problem(extraction_artifacts, "extraction", "存在重复 extractor", action="每个 extractor 只保留一份")
            )
            return _problems(problems)
        source_artifacts = {payload["source_artifact"] for payload in payloads}
        if len(source_artifacts) != 1:
            problems.append(
                _problem(
                    extraction_artifacts,
                    "extraction",
                    "抽取 artifact 必须指向同一 source_artifact",
                    action="确认所有 extractor 针对同一 PDF 结果",
                )
            )
            return _problems(problems)
    source_artifact = next(iter({payload["source_artifact"] for payload in payloads}), None)

    decision_map = {item["conflict_id"]: item["value"] for item in decisions}
    if payloads:
        initial_ab = set(extractor_ids) <= {"tecan-extractor-a", "tecan-extractor-b"}
        if initial_ab and decision_map:
            problems.append(
                _problem(
                    extraction_artifacts,
                    "decisions",
                    "C 抽取尚未完成就提供裁决",
                    action="先补 extractor C，再处理冲突",
                )
            )
            return _problems(problems)
        if initial_ab and len(payloads) < 2:
            problems.append(
                _problem(
                    source_artifact or extraction_artifacts,
                    "extraction",
                    "仅一个 A/B extractor 成功，需要第三个 extractor C",
                    action="主智能体补跑 extractor C",
                )
            )
            return _problems(problems)

    if problems or not payloads:
        return _problems(problems) if problems else _problems(
            [_problem(extraction_artifacts, "extraction", "没有可用的抽取 artifact", "重新发起抽取")]
        )

    used_decisions: set[str] = set()
    conflicts: list[dict[str, Any]] = []
    manual_checks: list[dict[str, Any]] = []
    missing: list[str] = []
    logistics: dict[str, Any] = {}
    for field_name in LOGISTICS_FIELDS:
        value = _resolve_vote(
            f"logistics.{field_name}",
            field_name,
            [(payload["extractor"], payload["logistics"][field_name]) for payload in payloads],
            decision_map,
            used_decisions,
            conflicts,
            manual_checks,
        )
        logistics[field_name] = value
        if value is None:
            missing.append(f"logistics.{field_name}")

    # A/B 阶段：任何冲突或缺失都先要求 extractor C，而不是直接裁决。
    if initial_ab and (conflicts or missing):
        problems.append(
            _problem(
                source_artifact or extraction_artifacts,
                "extraction",
                "A/B 抽取存在冲突或缺失，需要第三个 extractor C",
                action="主智能体补跑 extractor C 后再调用 generate_tecan_import",
            )
        )
        return _problems(problems)

    if conflicts:
        unknown = sorted(set(decision_map) - used_decisions)
        if unknown:
            problems.append(
                _problem(
                    extraction_artifacts,
                    "decisions",
                    f"未知 conflict_id: {', '.join(unknown)}",
                    action="只对实际冲突提交裁决",
                )
            )
            return _problems(problems)
        for conflict in conflicts:
            problems.append(
                _problem(
                    source_artifact or extraction_artifacts,
                    conflict["conflict_id"],
                    "抽取器对字段取值不一致",
                    action="回查 PDF，提交 conflict_id/value/reason 裁决",
                )
            )
        return _problems(problems)
    if set(decision_map) - used_decisions:
        problems.append(
            _problem(extraction_artifacts, "decisions", "裁决包含不存在的 conflict_id", action="只对实际冲突提交裁决")
        )
        return _problems(problems)
    if missing:
        for field in missing:
            problems.append(
                _problem(
                    source_artifact or extraction_artifacts,
                    field,
                    "所有抽取器均缺失该物流字段",
                    action="用户补正材料后重新发起",
                )
            )
        return _problems(problems)

    order = _parse_order_workbook(order_path, order_artifact)
    if order.get("code") == "input_problems":
        return order
    info_records = _parse_information_workbooks(info_paths)
    if isinstance(info_records, dict) and info_records.get("code") == "input_problems":
        return info_records

    candidates_by_pn = info_records
    selected: dict[str, dict[str, Any]] = {}
    info_conflicts: list[dict[str, Any]] = []
    for row in order["rows"]:
        pn = row["pn"]
        candidates = candidates_by_pn.get(pn, [])
        if not candidates:
            problems.append(
                _problem(
                    information_artifacts,
                    f"information.{pn}",
                    "信息表未提供该料号记录",
                    action="补全信息表中该料号的数据",
                )
            )
            return _problems(problems)
        distinct = _distinct_info_records(candidates)
        if len(distinct) == 1:
            selected[pn] = distinct[0]
            continue
        # 不再有 info_source_preference / pn_info_source_overrides：来源冲突一律作为问题。
        info_conflicts.append(
            {
                "pn": pn,
                "candidates": [
                    {
                        "source_artifact": item["source_artifact"],
                        "source_sheet": item["source_sheet"],
                        "source_row": item["source_row"],
                        "description": item["description"],
                        "origin_country": item["origin_country"],
                        "unit_net_weight": item["unit_net_weight"],
                    }
                    for item in distinct
                ],
            }
        )
    if info_conflicts:
        for conflict in info_conflicts:
            problems.append(
                _problem(
                    information_artifacts,
                    f"information.{conflict['pn']}",
                    "信息表对该料号存在多套不一致记录",
                    action="清理信息表，使每个料号只保留一套确定数据",
                )
            )
        return _problems(problems)

    total_net = Decimal("0")
    items: list[dict[str, Any]] = []
    for row in order["rows"]:
        info = selected[row["pn"]]
        net_weight = _quantize(_decimal(info["unit_net_weight"]) * Decimal(row["quantity"]))
        total_net += net_weight
        if info["source_sheet"] != "Sheet1":
            _manual(manual_checks, f"items.{row['pn']}.source_sheet", f"使用 {info['source_sheet']} 数据")
        items.append(
            {
                "pn": row["pn"],
                "quantity": row["quantity"],
                "net_price": _number(row["net_price"]),
                "amount": _number(row["amount"]),
                "description": info["description"],
                "origin_country": info["origin_country"],
                "unit_net_weight": _number(_decimal(info["unit_net_weight"])),
                "net_weight": _number(net_weight),
                "gross_weight": None,
                "source_artifact": info["source_artifact"],
                "source_sheet": info["source_sheet"],
                "source_row": info["source_row"],
            }
        )
    if total_net <= 0:
        problems.append(
            _problem(order_artifact, "items", "净重合计非正", action="核对订单数量与信息表净重")
        )
        return _problems(problems)
    gross_total = _decimal(logistics["gross_weight"])
    allocated = Decimal("0")
    for index, item in enumerate(items):
        net = _decimal(item["net_weight"])
        gross = gross_total - allocated if index == len(items) - 1 else _quantize(net / total_net * gross_total)
        item["gross_weight"] = _number(gross)
        allocated += gross
    logistics["pieces"] = int(_decimal(logistics["pieces"]))
    logistics["gross_weight"] = _number(gross_total)
    logistics["net_weight"] = _number(total_net)

    canonical = {
        "workflow": WORKFLOW,
        "source_artifacts": {
            "extractions": extraction_artifacts,
            "mineru": source_artifact,
            "order": order_artifact,
            "information": information_artifacts,
            "adjudication": None,
        },
        "logistics": logistics,
        "items": items,
        "manual_checks": manual_checks,
        "currency": order["currency"],
    }
    _validate_canonical(canonical)
    path = write_json_artifact("tecan_canonical", canonical)
    return {"canonical": canonical, "canonical_artifact": path, "manual_checks": manual_checks}


def _problem(source: Any, location: str, issue: str, action: str) -> dict[str, Any]:
    return {
        "source": _source_text(source),
        "location": location,
        "issue": issue,
        "action": action,
    }


def _source_text(source: Any) -> str:
    if isinstance(source, str):
        return source
    if isinstance(source, (list, tuple)) and source:
        return ", ".join(str(item) for item in source)
    return str(source)


def _problems(problems: list[dict[str, Any]]) -> dict[str, Any]:
    return {"code": "input_problems", "problems": problems}


# ---------- 抽取/合同校验 ----------


def _validate_extraction(payload: Mapping[str, Any]) -> dict[str, Any]:
    if set(payload) != {"workflow", "extractor", "source_artifact", "logistics", "items"}:
        raise ValueError("Invalid Tecan extraction envelope")
    if payload.get("workflow") != WORKFLOW or payload.get("extractor") not in EXTRACTORS:
        raise ValueError("Invalid Tecan workflow or extractor")
    source = payload.get("source_artifact")
    if not isinstance(source, str) or not source.startswith("/artifacts/"):
        raise ValueError("source_artifact must be an explicit /artifacts/... path")
    logistics = payload.get("logistics")
    if not isinstance(logistics, Mapping) or set(logistics) != set(LOGISTICS_FIELDS):
        raise ValueError("Tecan logistics must contain pieces and gross_weight")
    normalized: dict[str, dict[str, Any]] = {}
    for field_name in LOGISTICS_FIELDS:
        field = logistics[field_name]
        if not isinstance(field, Mapping) or set(field) != {"value", "confidence"}:
            raise ValueError(f"logistics.{field_name} must contain value and confidence")
        confidence = field.get("confidence")
        if confidence not in CONFIDENCES or (field.get("value") is None and confidence != "low"):
            raise ValueError(f"Invalid confidence for logistics.{field_name}")
        normalized[field_name] = {"value": field.get("value"), "confidence": confidence}
    if payload.get("items") != []:
        raise ValueError("Tecan extraction items must be an empty list")
    return {
        "workflow": WORKFLOW,
        "extractor": payload["extractor"],
        "source_artifact": source,
        "logistics": normalized,
        "items": [],
    }


def _load_extractions(paths: Sequence[str]) -> tuple[list[dict[str, Any]], list[str]]:
    payloads: list[dict[str, Any]] = []
    invalid: list[str] = []
    for path in paths:
        try:
            payloads.append(_validate_extraction(read_json_artifact(path)))
        except (OSError, ValueError) as exc:
            invalid.append(f"{path}: {exc}")
    return payloads, invalid


def _validate_decisions(decisions: Any) -> list[dict[str, Any]]:
    if not isinstance(decisions, list):
        return []
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for decision in decisions:
        if not isinstance(decision, Mapping) or set(decision) != {"conflict_id", "value", "reason"}:
            raise ValueError("Each decision must contain conflict_id, value and reason")
        conflict_id = decision.get("conflict_id")
        reason = decision.get("reason")
        if not isinstance(conflict_id, str) or not conflict_id or conflict_id in seen:
            raise ValueError("conflict_id must be unique and non-empty")
        if not isinstance(reason, str) or not reason.strip():
            raise ValueError("decision reason must be non-empty")
        seen.add(conflict_id)
        result.append({"conflict_id": conflict_id, "value": decision.get("value"), "reason": reason.strip()})
    return result


def _validate_canonical(payload: Mapping[str, Any]) -> dict[str, Any]:
    if set(payload) != CANONICAL_KEYS or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Tecan canonical contract")
    source = payload.get("source_artifacts")
    if not isinstance(source, Mapping) or set(source) != {"extractions", "mineru", "order", "information", "adjudication"}:
        raise ValueError("Invalid Tecan canonical source_artifacts")
    logistics = payload.get("logistics")
    if not isinstance(logistics, Mapping) or set(logistics) != {"pieces", "gross_weight", "net_weight"}:
        raise ValueError("Invalid Tecan canonical logistics")
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise ValueError("Tecan canonical items must not be empty")
    if any(not isinstance(item, Mapping) or set(item) != ITEM_KEYS for item in items):
        raise ValueError("Invalid Tecan canonical item")
    if not isinstance(payload.get("manual_checks"), list) or not isinstance(payload.get("currency"), str):
        raise ValueError("Invalid Tecan canonical metadata")
    return dict(payload)


# ---------- 物流投票 ----------


def _resolve_vote(
    conflict_id: str,
    field_name: str,
    fields: Sequence[tuple[str, Mapping[str, Any]]],
    decisions: Mapping[str, Any],
    used_decisions: set[str],
    conflicts: list[dict[str, Any]],
    manual_checks: list[dict[str, Any]],
) -> Any:
    values: dict[str, list[tuple[str, Any]]] = {}
    for extractor, field in fields:
        value = field.get("value")
        token = _numeric_token(value)
        if token is not None:
            values.setdefault(token, []).append((extractor, value))
    if not values:
        return None
    winner = max(values.values(), key=len)
    if len(winner) >= 2:
        return _canonical_number(field_name, winner[0][1])
    if len(values) == 1:
        _manual(manual_checks, conflict_id, "仅一个 extractor 提供非空值")
        return _canonical_number(field_name, winner[0][1])
    if conflict_id in decisions:
        used_decisions.add(conflict_id)
        return _canonical_number(field_name, decisions[conflict_id])
    conflicts.append(
        {
            "conflict_id": conflict_id,
            "field": conflict_id,
            "values": [
                {"extractor": extractor, "value": value}
                for entries in values.values()
                for extractor, value in entries
            ],
        }
    )
    return None


# ---------- 订单/信息表解析 ----------


def _parse_order_workbook(path: Path, artifact: str) -> dict[str, Any]:
    if path.suffix.lower() != ".xlsx":
        return _problems([_problem(artifact, "order_artifact", "订单必须是 .xlsx", action="提供 .xlsx 订单")])
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError):
        return _problems([_problem(artifact, "order_artifact", "订单无法读取", action="提供有效的订单 Excel")])
    try:
        matches = []
        for sheet in workbook.worksheets:
            cells = next(sheet.iter_rows(min_row=1, max_row=1), ())
            headers = {_normalize_header(cell.value): index for index, cell in enumerate(cells)}
            if {"PN", "Order Qty", "Amount"} <= set(headers):
                matches.append((sheet, headers))
        if len(matches) != 1:
            return _problems([_problem(artifact, "order_artifact", "找不到唯一的订单工作表", action="确认订单表头 PN/Order Qty/Amount")])
        sheet, headers = matches[0]
        rows: list[dict[str, Any]] = []
        currencies: set[str] = set()
        for excel_row, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            pn = normalize_pn(cells[headers["PN"]].value)
            if not pn and all(cells[index].value in (None, "") for index in headers.values()):
                continue
            quantity = _positive_integer(cells[headers["Order Qty"]].value)
            amount_cell = cells[headers["Amount"]]
            amount = _money(amount_cell.value)
            currencies.update(_currencies(amount_cell.value, amount_cell.number_format))
            net_price = None
            if "Net Price" in headers:
                price_cell = cells[headers["Net Price"]]
                net_price = _money(price_cell.value)
                currencies.update(_currencies(price_cell.value, price_cell.number_format))
            if net_price is None and amount is not None and quantity:
                net_price = amount / Decimal(quantity)
            if not pn or quantity is None or amount is None or net_price is None:
                return _problems([_problem(artifact, f"order_row.{excel_row}", "订单行缺失关键字段", action="补全该行数据")])
            rows.append({"pn": pn, "quantity": quantity, "net_price": net_price, "amount": amount})
        if not rows:
            return _problems([_problem(artifact, "order_artifact", "订单没有数据行", action="提供包含数据的订单")])
        if len(currencies) > 1:
            return _problems(
                [_problem(artifact, "order_artifact", f"订单存在多种币种: {sorted(currencies)}", action="确保订单单一币种")]
            )
        return {"status": "ok", "rows": rows, "currency": next(iter(currencies), "RMB")}
    finally:
        workbook.close()


def _parse_information_workbooks(
    paths: Sequence[tuple[str, Path]],
) -> dict[str, list[dict[str, Any]]] | dict[str, Any]:
    records_by_pn: dict[str, list[dict[str, Any]]] = {}
    for artifact, path in paths:
        if path.suffix.lower() != ".xlsx":
            return _problems([_problem(artifact, "information_artifacts", "信息表必须是 .xlsx", action="提供 .xlsx 信息表")])
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError, ValueError):
            return _problems([_problem(artifact, "information_artifacts", "信息表无法读取", action="提供有效的信息表")])
        try:
            sheets: list[tuple[Any, dict[str, int], str]] = []
            for sheet in workbook.worksheets:
                values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = {_normalize_header(value): index for index, value in enumerate(values)}
                weight = next((name for name in ("净重", "参考净重") if name in headers), None)
                if {"料号", "英文品名", "原产国"} <= set(headers) and weight:
                    sheets.append((sheet, headers, weight))
            if not sheets:
                return _problems([_problem(artifact, "information_artifacts", "信息表缺少料号/英文品名/原产国/净重表头", action="确认信息表表头")])
            per_sheet: dict[str, dict[str, list[dict[str, Any]]]] = {}
            for sheet, headers, weight_header in sheets:
                sheet_records: dict[str, list[dict[str, Any]]] = {}
                for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    pn = normalize_pn(values[headers["料号"]] if headers["料号"] < len(values) else None)
                    if not pn:
                        continue
                    description = _text(values[headers["英文品名"]] if headers["英文品名"] < len(values) else None)
                    country = _text(values[headers["原产国"]] if headers["原产国"] < len(values) else None)
                    weight = _decimal(values[headers[weight_header]] if headers[weight_header] < len(values) else None)
                    if not description or not country or weight is None or weight <= 0:
                        continue
                    sheet_records.setdefault(pn, []).append(
                        {
                            "description": description,
                            "origin_country": country,
                            "unit_net_weight": _number(weight),
                            "source_artifact": artifact,
                            "source_sheet": sheet.title,
                            "source_row": row_number,
                        }
                    )
                per_sheet[sheet.title] = sheet_records
            all_pns = {pn for records in per_sheet.values() for pn in records}
            for pn in all_pns:
                primary = per_sheet.get("Sheet1", {}).get(pn, [])
                candidates = primary or [item for name, records in per_sheet.items() if name != "Sheet1" for item in records.get(pn, [])]
                records_by_pn.setdefault(pn, []).extend(candidates)
        finally:
            workbook.close()
    return records_by_pn


def _distinct_info_records(records: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    keys: set[tuple[str, str, str]] = set()
    for record in records:
        key = (
            record["description"].casefold(),
            record["origin_country"].casefold(),
            _numeric_token(record["unit_net_weight"]) or "",
        )
        if key not in keys:
            keys.add(key)
            result.append(record)
    return result


# ---------- 规范化 / 数值 helper ----------


def normalize_pn(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        if not value.is_finite():
            return ""
        value = value.quantize(Decimal("1")) if value == value.to_integral_value() else value
        return format(value, "f")
    text = re.sub(r"\s+", "", str(value).strip())
    match = re.fullmatch(r"([+-]?\d+)\.0+", text)
    return match.group(1) if match else text


def _manual(checks: list[dict[str, Any]], field: str, reason: str) -> None:
    if not any(item.get("field") == field and item.get("reason") == reason for item in checks):
        checks.append({"field": field, "reason": reason})


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _positive_integer(value: Any) -> int | None:
    number = _decimal(value)
    if number is None or number <= 0 or number != number.to_integral_value():
        return None
    return int(number)


def _money(value: Any) -> Decimal | None:
    number = _decimal(value)
    if number is not None:
        return number
    if not isinstance(value, str):
        return None
    candidate = "".join(value.split()).replace("，", ",").replace("．", ".")
    candidate = _CURRENCY_RE.sub("", candidate).replace(",", "")
    try:
        result = Decimal(candidate)
        return result if result.is_finite() else None
    except InvalidOperation:
        return None


def _currencies(value: Any, number_format: str | None) -> set[str]:
    result: set[str] = set()
    text = f"{value if isinstance(value, str) else ''} {number_format or ''}".upper()
    for token in _CURRENCY_RE.findall(text):
        normalized = token.upper()
        if normalized in {"¥", "￥", "RMB", "CNY"}:
            result.add("RMB")
        elif normalized in {"$", "US$", "USD"}:
            result.add("USD")
        elif normalized in {"€", "EUR"}:
            result.add("EUR")
        elif normalized in {"£", "GBP"}:
            result.add("GBP")
        else:
            result.add(normalized)
    return result


def _canonical_number(field: str, value: Any) -> Any:
    number = _decimal(value)
    if number is None or number <= 0:
        return None
    if field == "pieces":
        return int(number) if number == number.to_integral_value() else None
    return _number(number)


def _numeric_token(value: Any) -> str | None:
    number = _decimal(value)
    return format(number.normalize(), "f") if number is not None and number > 0 else None


def _decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return Decimal(str(value))
    text = str(value).strip()
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
    if not match:
        return None
    try:
        result = Decimal(match.group(0).replace(",", ""))
        return result if result.is_finite() else None
    except InvalidOperation:
        return None


def _number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _quantize(value: Decimal) -> Decimal:
    return value.quantize(_WEIGHT_QUANT, rounding=ROUND_HALF_UP)


def _text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None
