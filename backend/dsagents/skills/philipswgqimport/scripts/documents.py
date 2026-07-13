"""Philips 外高桥进境确定性 Excel 写入：tracking、invoice/packing、核注清单。

只消费已经过校验的 canonical dict；不做业务判断、不做抽取。
共享的 openpyxl 列/行/样式 helper 也定义在这里，被 tools.py 复用。
"""
from __future__ import annotations

import copy
import re
from datetime import datetime, time
from decimal import ROUND_CEILING, Decimal
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from dsagents.integrations.artifacts import (
    resolve_artifact_path,
    to_virtual_artifact_path,
    unique_download_path,
)


_TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "assets"
_INVOICE_TEMPLATE = _TEMPLATE_DIR / "invoice,packing进境.xlsx"
_BONDED_TEMPLATE = _TEMPLATE_DIR / "核注清单导入模板.xlsx"


def generate_tracking(canonical: Mapping[str, Any]) -> str:
    source = resolve_artifact_path(canonical["source_artifacts"]["tracking"])
    workbook = load_workbook(source, data_only=False, read_only=False, keep_vba=source.suffix.lower() == ".xlsm")
    try:
        if "进口" not in workbook.sheetnames:
            raise ValueError("tracking workbook is missing 进口 sheet")
        sheet = workbook["进口"]
        headers = header_columns(sheet)
        required = {
            "product_id": ("料号",), "quantity": ("数量",), "description": ("型号",),
            "currency": ("币种",), "unit_price": ("备案单价USD", "备案单价"),
            "total_price": ("备案总价USD", "备案总价"), "origin_country": ("原产国",),
            "hawb": ("运单号",), "forwarder": ("国际货代",), "po": ("进境STO(PO)", "进境STO"),
            "pieces": ("件数",), "net": ("净重",), "gross": ("毛重",),
            "date": ("Figo提供完整申报要素日期",),
        }
        columns = {name: required_column(headers, aliases, name) for name, aliases in required.items()}
        remark_column = optional_column(headers, ("进境备注",))
        declaration_columns = declaration_header_columns(headers)
        last_row = last_data_row(sheet)
        max_column = max(sheet.max_column, max(columns.values()))
        declaration_date = datetime.combine(datetime.now().date(), time.min)
        for index, item in enumerate(canonical["items"], start=1):
            source_row = item.get("tracking_source_row") or last_row
            output_row = last_row + index
            copy_sheet_row(sheet, source_row, output_row, max_column)
            values = {
                "product_id": item["normalized_product_id"], "quantity": item["quantity"],
                "description": item["description"], "currency": item["currency"],
                "unit_price": item["unit_price"], "total_price": item["total_price"],
                "origin_country": item["origin_country"], "hawb": canonical["logistics"]["hawb_number"],
                "forwarder": canonical["international_forwarder"], "po": item["po_number"],
                "pieces": item["quantity"], "net": item["net_weight"], "gross": item["gross_weight"],
                "date": declaration_date,
            }
            for field, value in values.items():
                sheet.cell(output_row, columns[field]).value = value
            sheet.cell(output_row, 1).value = None
            if remark_column:
                sheet.cell(output_row, remark_column).value = "进境"
            for field_name, column in declaration_columns.items():
                if column:
                    sheet.cell(output_row, column).value = item["declaration"][field_name]
        target = unique_download_path(f"{source.stem}_进境更新", source.suffix)
        workbook.save(target)
    finally:
        workbook.close()
    return to_virtual_artifact_path(target)


def generate_invoice_packing(canonical: Mapping[str, Any]) -> str:
    workbook = load_workbook(_INVOICE_TEMPLATE, data_only=False)
    try:
        customs = workbook["Customs invoice"]
        packing = workbook["Packing List"]
        now = datetime.now()
        customs["F2"] = now.date()
        packing["F3"] = now.date()
        customs_total = prepare_detail_rows(customs, 32, 39, len(canonical["items"]), 11)
        packing_total = prepare_detail_rows(packing, 37, 44, len(canonical["items"]), 11)
        currency = canonical["items"][0]["currency"] or "需确认"
        total_amount = Decimal("0")
        for index, item in enumerate(canonical["items"]):
            row = 32 + index
            customs.cell(row, 1).value = index + 1
            customs.cell(row, 2).value = item["normalized_product_id"]
            customs.cell(row, 3).value = item["description"] or "需确认：描述"
            customs.cell(row, 6).value = item["quantity"]
            customs.cell(row, 8).value = item["raw_country"] or "需确认：原产国"
            customs.cell(row, 9).value = item["unit_price"]
            customs.cell(row, 11).value = item["total_price"]
            total_amount += to_decimal(item["total_price"]) or Decimal("0")

            row = 37 + index
            packing.cell(row, 1).value = index + 1
            packing.cell(row, 2).value = item["normalized_product_id"]
            packing.cell(row, 3).value = item["description"] or "需确认：描述"
            packing.cell(row, 6).value = item["quantity"]
            packing.cell(row, 7).value = item["net_weight"]
            packing.cell(row, 10).value = item["gross_weight"]
        customs.cell(25, 9).value = f"( {currency})"
        customs.cell(25, 11).value = f"({currency})"
        customs.cell(customs_total, 4).value = "Total Amount:"
        customs.cell(customs_total, 9).value = currency
        customs.cell(customs_total, 11).value = to_number(total_amount)
        packing.cell(packing_total, 4).value = "TOTAL PCS:"
        packing.cell(packing_total, 6).value = to_number(
            sum_decimal(item["quantity"] for item in canonical["items"])
        )
        packing.cell(packing_total + 1, 4).value = "TOTAL NET WEIGHT:"
        packing.cell(packing_total + 1, 6).value = canonical["logistics"]["net_weight"]
        packing.cell(packing_total + 2, 4).value = "TOTAL GROSS WEIGHT: "
        packing.cell(packing_total + 2, 6).value = canonical["logistics"]["gross_weight"]
        target = unique_download_path("invoice_packing进境", ".xlsx")
        workbook.save(target)
    finally:
        workbook.close()
    return to_virtual_artifact_path(target)


def generate_bonded_checklist(
    canonical: Mapping[str, Any],
    units: Mapping[str, tuple[str, str, str]],
) -> str:
    workbook = load_workbook(_BONDED_TEMPLATE, data_only=False)
    try:
        header = workbook["表头"]
        body = workbook["表体"]
        header["M2"] = f"{datetime.now():%Y%m%d}"
        header["S2"] = canonical["logistics"]["shipper_country_name"]
        header["T2"] = "2244" if canonical["customs_mode"] == "快件" else "2233"
        prepare_bonded_rows(body, len(canonical["items"]))
        currency_names = {"USD": "美元", "CNY": "人民币", "RMB": "人民币", "EUR": "欧元"}
        for index, item in enumerate(canonical["items"]):
            row = 2 + index
            declaration = item["declaration"]
            declaration_unit, legal_unit, legal_second = units[item["normalized_product_id"]]
            values = {
                2: index + 1, 4: index + 1, 6: item["normalized_product_id"],
                7: declaration["hs_code"], 8: declaration["chinese_name"], 9: item["description"],
                10: declaration_unit, 11: item["quantity"], 12: legal_unit, 13: item["quantity"],
                14: legal_second, 15: item["net_weight"], 16: item["unit_price"],
                17: item["total_price"], 18: item["origin_country"],
                19: currency_names.get(item["currency"], item["currency"]),
                20: item["gross_weight"], 21: item["net_weight"],
                32: "1" if item["raw_country"] == "US" else "2",
            }
            for column, value in values.items():
                body.cell(row, column).value = value
        target = unique_download_path("核注清单导入模板_进境", ".xlsx")
        workbook.save(target)
    finally:
        workbook.close()
    return to_virtual_artifact_path(target)


# ---------- 共享 openpyxl helper（也被 tools.py 的 tracking 查询复用） ----------


def header_columns(sheet: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        header = re.sub(r"\s+", "", str(sheet.cell(1, column).value or "")).lower()
        if header and header not in result:
            result[header] = column
    return result


def optional_column(headers: Mapping[str, int], aliases: Sequence[str]) -> int | None:
    for alias in aliases:
        column = headers.get(re.sub(r"\s+", "", alias).lower())
        if column:
            return column
    return None


def required_column(headers: Mapping[str, int], aliases: Sequence[str], field: str) -> int:
    column = optional_column(headers, aliases)
    if column is None:
        raise ValueError(f"tracking workbook missing required column: {field}")
    return column


def last_data_row(sheet: Any) -> int:
    for row in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row, column).value not in (None, "") for column in range(1, sheet.max_column + 1)):
            return row
    return 2


def prepare_detail_rows(sheet: Any, start: int, total: int, count: int, max_column: int) -> int:
    if count <= 1:
        return total
    sheet.insert_rows(start + 1, amount=count - 1)
    for row in range(start + 1, start + count):
        copy_style_row(sheet, start, row, max_column, copy_values=False)
    return total + count - 1


def prepare_bonded_rows(sheet: Any, count: int) -> None:
    if count <= 1:
        return
    sheet.insert_rows(3, amount=count - 1)
    for row in range(3, 2 + count):
        copy_style_row(sheet, 2, row, 32, copy_values=True)


def copy_sheet_row(sheet: Any, source_row: int, target_row: int, max_column: int) -> None:
    source_row = max(2, min(source_row, sheet.max_row))
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        value = source.value
        if isinstance(value, str) and value.startswith("="):
            try:
                value = Translator(value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                pass
        target.value = value
        copy_cell_style(source, target)


def copy_style_row(sheet: Any, source_row: int, target_row: int, max_column: int, *, copy_values: bool) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target.value = source.value if copy_values else None
        copy_cell_style(source, target)


def copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
    target.number_format = source.number_format


def declaration_header_columns(headers: Mapping[str, int]) -> dict[str, int | None]:
    from dsagents.skills.philipswgqimport.scripts.tools import DECLARATION_HEADER_ALIASES

    return {
        field: optional_column(headers, aliases)
        for field, aliases in DECLARATION_HEADER_ALIASES.items()
    }


# ---------- decimal/number helper（写 Excel 的数值规范化用） ----------


def to_decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, (int, float)):
        try:
            result = Decimal(str(value))
            return result if result.is_finite() else None
        except Exception:
            return None
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", str(value).strip())
    if not match:
        return None
    try:
        result = Decimal(match.group(0).replace(",", ""))
        return result if result.is_finite() else None
    except Exception:
        return None


def to_number(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    return int(value) if value == value.to_integral_value() else float(value)


def sum_decimal(values: Any, *, allow_missing: bool = False) -> Decimal | None:
    result = Decimal("0")
    found = False
    for value in values:
        number = to_decimal(value)
        if number is None:
            if allow_missing:
                continue
            return None
        found = True
        result += number
    return result if found else None


def allocate_total(total: Decimal, weights: Sequence[Decimal]) -> list[Decimal]:
    if not weights:
        return []
    weight_total = sum(weights, Decimal("0"))
    if weight_total <= 0:
        weights = [Decimal("1") for _ in weights]
        weight_total = Decimal(len(weights))
    quant = Decimal(1).scaleb(min(total.as_tuple().exponent, 0))
    remaining = total
    result: list[Decimal] = []
    for index, weight in enumerate(weights):
        if index == len(weights) - 1:
            share = remaining
        else:
            share = (total * weight / weight_total).quantize(quant, rounding=ROUND_CEILING)
        share = min(max(share, Decimal("0")), remaining)
        result.append(share)
        remaining -= share
    return result
