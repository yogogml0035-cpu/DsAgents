from __future__ import annotations

import copy
import os
import re
from datetime import datetime, time
from decimal import ROUND_CEILING, Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from workflow_artifacts import (
    read_json_artifact,
    resolve_artifact_path,
    to_virtual_artifact_path,
    unique_download_path,
    write_json_artifact,
)


WORKFLOW = "philips-wgq-import"
EXTRACTORS = frozenset(f"philips-wgq-extractor-{name}" for name in "abc")
CONFIDENCES = frozenset({"high", "medium", "low"})
LOGISTICS_FIELDS = ("hawb_number", "pieces", "gross_weight", "shipper_country")
ITEM_FIELDS = (
    "product_id_raw",
    "description",
    "quantity",
    "unit_price",
    "total_price",
    "currency",
    "po_number",
    "raw_country",
    "gross_weight",
)
VALID_FORWARDERS = frozenset({"DHL", "DSV", "FEDEX", "UPS", "康捷空"})
COUNTRY_NAMES = {
    "BE": "比利时", "CA": "加拿大", "CH": "瑞士", "CN": "中国",
    "DE": "德国", "FR": "法国", "GB": "英国", "HK": "中国香港",
    "IT": "意大利", "JP": "日本", "NL": "荷兰", "SG": "新加坡", "US": "美国",
}
COUNTRY_ALIASES = {
    "BELGIUM": "BE", "比利时": "BE", "CANADA": "CA", "加拿大": "CA",
    "SWITZERLAND": "CH", "瑞士": "CH", "CHINA": "CN", "中国": "CN",
    "GERMANY": "DE", "德国": "DE", "FRANCE": "FR", "法国": "FR",
    "UNITED KINGDOM": "GB", "UK": "GB", "英国": "GB", "HONG KONG": "HK",
    "HONG KONG CHINA": "HK",
    "中国香港": "HK", "ITALY": "IT", "意大利": "IT", "JAPAN": "JP",
    "日本": "JP", "NETHERLANDS": "NL", "THE NETHERLANDS": "NL", "荷兰": "NL",
    "SINGAPORE": "SG", "新加坡": "SG", "UNITED STATES": "US",
    "UNITED STATES OF AMERICA": "US", "USA": "US", "美国": "US",
}
DECLARATION_HEADER_ALIASES = {
    "hs_code": ("HS code", "HS编码", "HScode", "商品编码"),
    "chinese_name": ("中文品名",),
    "legal_unit": ("法定第一单位", "法定单位", "申报单位"),
    "modality": ("Modality",),
    "regulatory_condition": ("监管条件",),
    "regulatory_condition_details": ("监管条件详解", "监管条件详细"),
    "customs_commodity_name": ("海关税则书所列商品名称",),
}
DECLARATION_FIELDS = tuple(DECLARATION_HEADER_ALIASES)
CANONICAL_KEYS = frozenset(
    {
        "workflow",
        "source_artifacts",
        "logistics",
        "items",
        "manual_checks",
        "international_forwarder",
        "customs_mode",
    }
)
CANONICAL_ITEM_KEYS = frozenset(
    {
        "product_id_raw",
        "product_id",
        "normalized_product_id",
        "description",
        "quantity",
        "unit_price",
        "total_price",
        "currency",
        "po_number",
        "raw_country",
        "origin_country",
        "gross_weight",
        "net_weight",
        "tracking_source_row",
        "declaration",
    }
)
_INVOICE_TEMPLATE = Path(__file__).resolve().parent / "skills" / WORKFLOW / "assets" / "invoice,packing进境.xlsx"
_BONDED_TEMPLATE = Path(__file__).resolve().parent / "skills" / WORKFLOW / "assets" / "核注清单导入模板.xlsx"
_ORACLE_SQL = """
select c.jldw, u1.unit_name, u2.unit_name
from od.chda c
left join dongsong.good g on g.goodcode = '01' || c.hsbm
left join dongsong.custom_unit u1 on u1.unit_code = g.unit
left join dongsong.custom_unit u2 on u2.unit_code = g.unit2
where c.chbm = :product_id
"""
_ORACLE_CLIENT_INITIALIZED = False


def save_philips_wgq_extraction(
    extractor: str,
    source_artifact: str,
    logistics: dict[str, dict[str, Any]],
    items: list[dict[str, dict[str, Any]]],
) -> dict[str, str]:
    """Validate and immutably save one Philips extraction artifact."""
    if not resolve_artifact_path(source_artifact).is_file():
        raise ValueError(f"source_artifact not found: {source_artifact}")
    payload = _validate_extraction(
        {
            "workflow": WORKFLOW,
            "extractor": extractor,
            "source_artifact": source_artifact,
            "logistics": logistics,
            "items": items,
        }
    )
    path = write_json_artifact(extractor, payload)
    return {"extractor": extractor, "artifact_path": path}


def save_philips_wgq_adjudication(
    source_artifacts: list[str],
    decisions: list[dict[str, Any]],
) -> dict[str, str]:
    """Save only explicit conflict decisions; canonical validation happens later."""
    if not source_artifacts:
        raise ValueError("source_artifacts must not be empty")
    for path in source_artifacts:
        if not resolve_artifact_path(path).is_file():
            raise ValueError(f"source_artifact not found: {path}")
    normalized = _validate_decisions(decisions)
    path = write_json_artifact(
        "philips_wgq_adjudication",
        {"workflow": WORKFLOW, "source_artifacts": source_artifacts, "decisions": normalized},
    )
    return {"artifact_path": path}


def build_philips_wgq_canonical(
    extraction_artifacts: list[str],
    tracking_artifact: str,
    international_forwarder: str | None = None,
    customs_mode: str = "普货",
    adjudication_artifact: str | None = None,
) -> dict[str, Any]:
    """Reconcile explicit A/B/C artifacts and write one strict canonical artifact."""
    forwarder = _normalize_forwarder(international_forwarder)
    if forwarder is None:
        return {"status": "needs_input", "missing": ["international_forwarder"]}
    mode = "快件" if "快件" in str(customs_mode) else "普货"
    tracking_path = resolve_artifact_path(tracking_artifact)
    if tracking_path.suffix.lower() not in {".xlsx", ".xlsm"} or not tracking_path.is_file():
        return {"status": "needs_input", "missing": ["tracking_artifact"]}

    payloads, invalid = _load_extractions(extraction_artifacts)
    if not payloads:
        return {"status": "needs_input", "missing": ["valid_extraction"]}
    extractor_ids = {payload["extractor"] for payload in payloads}
    if len(extractor_ids) != len(payloads):
        raise ValueError("Duplicate extractor artifacts are not allowed")
    source_artifacts = {payload["source_artifact"] for payload in payloads}
    if len(source_artifacts) != 1:
        raise ValueError("All extraction artifacts must reference the same source_artifact")
    source_artifact = next(iter(source_artifacts))

    initial_ab = extractor_ids <= {"philips-wgq-extractor-a", "philips-wgq-extractor-b"}
    if initial_ab and adjudication_artifact is not None:
        raise ValueError("Extractor C is required before adjudication")
    if initial_ab and len(payloads) < 2:
        return _needs_c(source_artifact, ["only_one_extractor_succeeded", *invalid])

    decision_payload = _load_adjudication(adjudication_artifact) if adjudication_artifact else None
    if decision_payload and decision_payload["source_artifacts"] != extraction_artifacts:
        raise ValueError("Adjudication source_artifacts do not match extraction_artifacts")
    decisions = {
        item["conflict_id"]: item["value"]
        for item in (decision_payload or {}).get("decisions", [])
    }
    manual_checks: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    used_decisions: set[str] = set()
    jointly_missing: list[str] = []

    logistics: dict[str, Any] = {}
    for field_name in LOGISTICS_FIELDS:
        value = _resolve_vote(
            f"logistics.{field_name}",
            field_name,
            [(payload["extractor"], payload["logistics"][field_name]) for payload in payloads],
            decisions,
            used_decisions,
            conflicts,
            manual_checks,
        )
        logistics[field_name] = value
        if value is None:
            jointly_missing.append(f"logistics.{field_name}")

    resolved_rows: list[dict[str, Any]] = []
    for row_key, row_fields in _aligned_items(payloads):
        row: dict[str, Any] = {}
        for field_name in ITEM_FIELDS:
            row[field_name] = _resolve_vote(
                f"items.{row_key}.{field_name}",
                field_name,
                row_fields[field_name],
                decisions,
                used_decisions,
                conflicts,
                manual_checks,
            )
        if row["product_id_raw"] is None:
            jointly_missing.append(f"items.{row_key}.product_id_raw")
        if _positive_decimal(row["quantity"]) is None:
            jointly_missing.append(f"items.{row_key}.quantity")
        if row["product_id_raw"] is not None and _positive_decimal(row["quantity"]) is not None:
            resolved_rows.append(row)

    if initial_ab and (conflicts or jointly_missing or not resolved_rows):
        reasons = [item["field"] for item in conflicts] + jointly_missing
        if not resolved_rows:
            reasons.append("valid_items")
        return _needs_c(source_artifact, reasons)
    if conflicts:
        unknown = sorted(set(decisions) - used_decisions)
        if unknown:
            raise ValueError("Unknown conflict_id: " + ", ".join(unknown))
        return {
            "status": "needs_adjudication",
            "source_artifact": source_artifact,
            "conflicts": conflicts,
        }
    if set(decisions) - used_decisions:
        raise ValueError("Adjudication contains conflict_id values that are not present")
    if jointly_missing:
        return {"status": "needs_input", "missing": jointly_missing}
    if not resolved_rows:
        return {"status": "needs_input", "missing": ["valid_items"]}

    items, merge_error = _merge_items(resolved_rows, manual_checks)
    if merge_error:
        return {"status": "needs_input", **merge_error}
    tracking = _tracking_context(tracking_path, [item["normalized_product_id"] for item in items])
    _apply_weight_history(items, logistics["gross_weight"], tracking, manual_checks)
    for item in items:
        context = tracking.get(item["normalized_product_id"], {})
        item["tracking_source_row"] = context.get("source_row")
        item["declaration"] = context.get("declaration", _empty_declaration())
        for field_name, value in item["declaration"].items():
            if value == f"需确认：{field_name}":
                _manual(
                    manual_checks,
                    f"items.{item['normalized_product_id']}.declaration.{field_name}",
                    "tracking 未提供申报要素",
                )

    shipper_code = _country_code(logistics["shipper_country"])
    logistics["shipper_country"] = shipper_code or logistics["shipper_country"]
    logistics["shipper_country_name"] = COUNTRY_NAMES.get(shipper_code or "", logistics["shipper_country"])
    net_weight = _sum_decimal(item["net_weight"] for item in items)
    logistics["net_weight"] = _number(net_weight) if net_weight is not None else "未找到：净重"

    canonical = {
        "workflow": WORKFLOW,
        "source_artifacts": {
            "extractions": extraction_artifacts,
            "mineru": source_artifact,
            "tracking": tracking_artifact,
            "adjudication": adjudication_artifact,
        },
        "logistics": logistics,
        "items": items,
        "manual_checks": manual_checks,
        "international_forwarder": forwarder,
        "customs_mode": mode,
    }
    _validate_canonical(canonical)
    path = write_json_artifact("philips_wgq_canonical", canonical)
    return {"status": "canonical", "canonical_artifact": path, "manual_checks": manual_checks}


def generate_philips_wgq_documents(canonical_artifact: str) -> dict[str, Any]:
    """Generate the three Philips workbooks from one canonical artifact path."""
    canonical = _validate_canonical(read_json_artifact(canonical_artifact))
    units, oracle_checks = _oracle_units(canonical["items"])
    outputs = [
        _generate_tracking(canonical),
        _generate_invoice_packing(canonical),
        _generate_bonded_checklist(canonical, units),
    ]
    return {
        "status": "generated",
        "canonical_artifact": canonical_artifact,
        "artifacts": outputs,
        "manual_checks": [*canonical["manual_checks"], *oracle_checks],
    }


def _validate_extraction(payload: Mapping[str, Any]) -> dict[str, Any]:
    expected = {"workflow", "extractor", "source_artifact", "logistics", "items"}
    if set(payload) != expected or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Philips extraction envelope")
    extractor = payload.get("extractor")
    if extractor not in EXTRACTORS:
        raise ValueError("Unsupported Philips extractor")
    source = payload.get("source_artifact")
    if not isinstance(source, str) or not source.startswith("/artifacts/"):
        raise ValueError("source_artifact must be an explicit /artifacts/... path")
    logistics = _validate_field_map(payload.get("logistics"), LOGISTICS_FIELDS, "logistics")
    item_values = payload.get("items")
    if not isinstance(item_values, list):
        raise ValueError("items must be a list")
    items = [
        _validate_field_map(item, ITEM_FIELDS, f"items[{index}]")
        for index, item in enumerate(item_values)
    ]
    return {
        "workflow": WORKFLOW,
        "extractor": extractor,
        "source_artifact": source,
        "logistics": logistics,
        "items": items,
    }


def _validate_field_map(value: Any, fields: Sequence[str], label: str) -> dict[str, dict[str, Any]]:
    if not isinstance(value, Mapping) or set(value) != set(fields):
        raise ValueError(f"{label} must contain exactly: {', '.join(fields)}")
    result: dict[str, dict[str, Any]] = {}
    for field_name in fields:
        field = value[field_name]
        if not isinstance(field, Mapping) or set(field) != {"value", "confidence"}:
            raise ValueError(f"{label}.{field_name} must contain value and confidence")
        confidence = field.get("confidence")
        if confidence not in CONFIDENCES or (field.get("value") is None and confidence != "low"):
            raise ValueError(f"Invalid confidence for {label}.{field_name}")
        result[field_name] = {"value": field.get("value"), "confidence": confidence}
    return result


def _load_extractions(paths: Sequence[str]) -> tuple[list[dict[str, Any]], list[str]]:
    payloads: list[dict[str, Any]] = []
    invalid: list[str] = []
    for path in paths:
        try:
            payloads.append(_validate_extraction(read_json_artifact(path)))
        except (OSError, ValueError) as exc:
            invalid.append(f"{path}: {exc}")
    return payloads, invalid


def _load_adjudication(path: str) -> dict[str, Any]:
    payload = read_json_artifact(path)
    if set(payload) != {"workflow", "source_artifacts", "decisions"} or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Philips adjudication artifact")
    if not isinstance(payload.get("source_artifacts"), list):
        raise ValueError("adjudication source_artifacts must be a list")
    return {**payload, "decisions": _validate_decisions(payload.get("decisions"))}


def _validate_decisions(decisions: Any) -> list[dict[str, Any]]:
    if not isinstance(decisions, list):
        raise ValueError("decisions must be a list")
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for decision in decisions:
        if not isinstance(decision, Mapping) or set(decision) != {"conflict_id", "value", "reason"}:
            raise ValueError("Each decision must contain conflict_id, value and reason")
        conflict_id = decision.get("conflict_id")
        reason = decision.get("reason")
        if not isinstance(conflict_id, str) or not conflict_id or conflict_id in seen:
            raise ValueError("conflict_id must be unique and non-empty")
        if not isinstance(reason, str) or not reason.strip():
            raise ValueError("decision reason must be non-empty")
        seen.add(conflict_id)
        result.append({"conflict_id": conflict_id, "value": decision.get("value"), "reason": reason.strip()})
    return result


def _aligned_items(payloads: Sequence[dict[str, Any]]) -> list[tuple[str, dict[str, list[tuple[str, dict[str, Any]]]]]]:
    rows = []
    for index in range(max((len(payload["items"]) for payload in payloads), default=0)):
        aligned = {field: [] for field in ITEM_FIELDS}
        for payload in payloads:
            if index >= len(payload["items"]):
                continue
            item = payload["items"][index]
            for field_name in ITEM_FIELDS:
                aligned[field_name].append((payload["extractor"], item[field_name]))
        rows.append((f"row-{index + 1}", aligned))
    return rows


def _resolve_vote(
    conflict_id: str,
    field_name: str,
    fields: Sequence[tuple[str, Mapping[str, Any]]],
    decisions: Mapping[str, Any],
    used_decisions: set[str],
    conflicts: list[dict[str, Any]],
    manual_checks: list[dict[str, Any]],
) -> Any:
    values_by_token: dict[str, list[tuple[str, Any]]] = {}
    for extractor, field in fields:
        value = field.get("value")
        token = _comparison_token(field_name, value)
        if token is not None:
            values_by_token.setdefault(token, []).append((extractor, value))
    if not values_by_token:
        return None
    winner = max(values_by_token.values(), key=len)
    if len(winner) >= 2:
        return _canonical_value(field_name, winner[0][1])
    if len(values_by_token) == 1:
        _manual(manual_checks, conflict_id, "仅一个 extractor 提供非空值")
        return _canonical_value(field_name, winner[0][1])
    if conflict_id in decisions:
        used_decisions.add(conflict_id)
        return _canonical_value(field_name, decisions[conflict_id])
    conflicts.append(
        {
            "conflict_id": conflict_id,
            "field": conflict_id,
            "values": [
                {"extractor": extractor, "value": value}
                for entries in values_by_token.values()
                for extractor, value in entries
            ],
        }
    )
    return None


def _merge_items(
    rows: Sequence[dict[str, Any]],
    manual_checks: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    order: list[str] = []
    for row in rows:
        product_id = clean_product_id(row["product_id_raw"])
        normalized = normalize_product_id(product_id)
        if not normalized:
            continue
        if normalized not in grouped:
            grouped[normalized] = []
            order.append(normalized)
        grouped[normalized].append(row)

    result: list[dict[str, Any]] = []
    for normalized in order:
        group = grouped[normalized]
        unit_prices = {_decimal_token(row["unit_price"]) for row in group if _decimal(row["unit_price"]) is not None}
        if len(unit_prices) > 1:
            return [], {"missing": [f"items.{normalized}.unit_price"], "reason": "unit_price_conflict"}
        quantity = _sum_decimal(row["quantity"] for row in group)
        total_price = _sum_decimal(row["total_price"] for row in group)
        unit_price = _first_nonempty(row["unit_price"] for row in group)
        if unit_price is None and quantity and total_price is not None:
            unit_price = total_price / quantity
            _manual(manual_checks, f"items.{normalized}.unit_price", "由总价/数量推导")
        item = {
            "product_id_raw": _first_nonempty(row["product_id_raw"] for row in group),
            "product_id": clean_product_id(_first_nonempty(row["product_id_raw"] for row in group)),
            "normalized_product_id": normalized,
            "description": _consistent_or_first(group, "description", manual_checks, normalized),
            "quantity": _number(quantity),
            "unit_price": _number(_decimal(unit_price)),
            "total_price": _number(total_price),
            "currency": _canonical_value("currency", _consistent_or_first(group, "currency", manual_checks, normalized)),
            "po_number": _po_numbers(row["po_number"] for row in group),
            "raw_country": _country_code(_consistent_or_first(group, "raw_country", manual_checks, normalized)),
            "origin_country": None,
            "gross_weight": _number(
                _sum_decimal((row["gross_weight"] for row in group), allow_missing=True)
            ),
            "net_weight": None,
            "tracking_source_row": None,
            "declaration": _empty_declaration(),
        }
        item["origin_country"] = COUNTRY_NAMES.get(item["raw_country"] or "", item["raw_country"])
        for required in (
            "description", "unit_price", "total_price", "currency", "po_number", "raw_country", "gross_weight"
        ):
            if item[required] is None:
                _manual(manual_checks, f"items.{normalized}.{required}", "字段缺失")
        result.append(item)
    return result, None


def _tracking_context(path: Path, product_ids: Sequence[str]) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        if "进口" not in workbook.sheetnames:
            return {}
        sheet = workbook["进口"]
        headers = _headers(sheet)
        product_column = _optional_column(headers, ("料号", "飞利浦料号", "Product ID"))
        quantity_column = _optional_column(headers, ("数量",))
        gross_column = _optional_column(headers, ("毛重",))
        net_column = _optional_column(headers, ("净重",))
        declarations = _declaration_records(workbook)
        fallback_row = max(2, _last_data_row(sheet))
        result: dict[str, dict[str, Any]] = {}
        for product_id in product_ids:
            rows = [] if product_column is None else [
                row
                for row in range(2, sheet.max_row + 1)
                if normalize_product_id(sheet.cell(row, product_column).value) == product_id
            ]
            source_row = rows[-1] if rows else fallback_row
            history = None
            if quantity_column and gross_column and net_column:
                for row in reversed(rows):
                    quantity = _positive_decimal(sheet.cell(row, quantity_column).value)
                    gross = _positive_decimal(sheet.cell(row, gross_column).value)
                    net = _positive_decimal(sheet.cell(row, net_column).value)
                    if quantity and gross and net and net <= gross:
                        history = {"quantity": quantity, "gross": gross, "net": net}
                        break
            result[product_id] = {
                "source_row": source_row,
                "history": history,
                "declaration": declarations.get(product_id, _empty_declaration()),
            }
        return result
    finally:
        workbook.close()


def _declaration_records(workbook: Any) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for sheet_name in ("进口", "申报要素"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        headers = _headers(sheet)
        product_column = _optional_column(headers, ("飞利浦料号", "料号", "Product ID"))
        if product_column is None:
            continue
        columns = {
            field: _optional_column(headers, aliases)
            for field, aliases in DECLARATION_HEADER_ALIASES.items()
        }
        for row in range(2, sheet.max_row + 1):
            product_id = normalize_product_id(sheet.cell(row, product_column).value)
            if not product_id:
                continue
            record = records.setdefault(product_id, _empty_declaration())
            for field, column in columns.items():
                value = sheet.cell(row, column).value if column else None
                if value not in (None, ""):
                    record[field] = value
    return records


def _apply_weight_history(
    items: list[dict[str, Any]],
    gross_total_value: Any,
    tracking: Mapping[str, Mapping[str, Any]],
    manual_checks: list[dict[str, Any]],
) -> None:
    gross_total = _decimal(gross_total_value)
    if gross_total is None or gross_total <= 0:
        return
    weights: list[Decimal] = []
    for item in items:
        extracted = _decimal(item["gross_weight"])
        history = tracking.get(item["normalized_product_id"], {}).get("history")
        quantity = _decimal(item["quantity"]) or Decimal("0")
        if extracted and extracted > 0:
            weights.append(extracted)
        elif history:
            weights.append(history["gross"] * quantity / history["quantity"])
        else:
            weights.append(quantity)
    allocations = _allocate_total(gross_total, weights)
    for item, gross in zip(items, allocations, strict=True):
        item["gross_weight"] = _number(gross)
        history = tracking.get(item["normalized_product_id"], {}).get("history")
        if not history:
            item["net_weight"] = "未找到：净重"
            _manual(manual_checks, f"items.{item['normalized_product_id']}.net_weight", "tracking 无有效历史净毛重")
            continue
        net = (gross * history["net"] / history["gross"]).to_integral_value(rounding=ROUND_CEILING)
        item["net_weight"] = int(max(Decimal("0"), net))


def _allocate_total(total: Decimal, weights: Sequence[Decimal]) -> list[Decimal]:
    if not weights:
        return []
    weight_total = sum(weights, Decimal("0"))
    if weight_total <= 0:
        weights = [Decimal("1") for _ in weights]
        weight_total = Decimal(len(weights))
    quant = Decimal(1).scaleb(min(total.as_tuple().exponent, 0))
    remaining = total
    result: list[Decimal] = []
    for index, weight in enumerate(weights):
        share = remaining if index == len(weights) - 1 else (total * weight / weight_total).quantize(quant, rounding=ROUND_CEILING)
        share = min(max(share, Decimal("0")), remaining)
        result.append(share)
        remaining -= share
    return result


def _validate_canonical(payload: Mapping[str, Any]) -> dict[str, Any]:
    if set(payload) != CANONICAL_KEYS or payload.get("workflow") != WORKFLOW:
        raise ValueError("Invalid Philips canonical contract")
    source = payload.get("source_artifacts")
    if not isinstance(source, Mapping) or set(source) != {"extractions", "mineru", "tracking", "adjudication"}:
        raise ValueError("Invalid Philips canonical source_artifacts")
    logistics = payload.get("logistics")
    if not isinstance(logistics, Mapping) or set(logistics) != {
        "hawb_number", "pieces", "gross_weight", "shipper_country", "shipper_country_name", "net_weight"
    }:
        raise ValueError("Invalid Philips canonical logistics")
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise ValueError("Philips canonical items must not be empty")
    for item in items:
        if not isinstance(item, Mapping) or set(item) != CANONICAL_ITEM_KEYS:
            raise ValueError("Invalid Philips canonical item")
        declaration = item.get("declaration")
        if not isinstance(declaration, Mapping) or set(declaration) != set(DECLARATION_FIELDS):
            raise ValueError("Invalid Philips canonical declaration")
    if not isinstance(payload.get("manual_checks"), list):
        raise ValueError("manual_checks must be a list")
    return dict(payload)


def _generate_tracking(canonical: Mapping[str, Any]) -> str:
    source = resolve_artifact_path(canonical["source_artifacts"]["tracking"])
    workbook = load_workbook(source, data_only=False, read_only=False, keep_vba=source.suffix.lower() == ".xlsm")
    if "进口" not in workbook.sheetnames:
        raise ValueError("tracking workbook is missing 进口 sheet")
    sheet = workbook["进口"]
    headers = _headers(sheet)
    required = {
        "product_id": ("料号",), "quantity": ("数量",), "description": ("型号",),
        "currency": ("币种",), "unit_price": ("备案单价USD", "备案单价"),
        "total_price": ("备案总价USD", "备案总价"), "origin_country": ("原产国",),
        "hawb": ("运单号",), "forwarder": ("国际货代",), "po": ("进境STO(PO)", "进境STO"),
        "pieces": ("件数",), "net": ("净重",), "gross": ("毛重",),
        "date": ("Figo提供完整申报要素日期",),
    }
    columns = {name: _required_column(headers, aliases, name) for name, aliases in required.items()}
    remark_column = _optional_column(headers, ("进境备注",))
    declaration_columns = {
        field: _optional_column(headers, aliases)
        for field, aliases in DECLARATION_HEADER_ALIASES.items()
    }
    last_row = _last_data_row(sheet)
    max_column = max(sheet.max_column, max(columns.values()))
    declaration_date = datetime.combine(datetime.now().date(), time.min)
    for index, item in enumerate(canonical["items"], start=1):
        source_row = item.get("tracking_source_row") or last_row
        output_row = last_row + index
        _copy_sheet_row(sheet, source_row, output_row, max_column)
        values = {
            "product_id": item["normalized_product_id"], "quantity": item["quantity"],
            "description": item["description"], "currency": item["currency"],
            "unit_price": item["unit_price"], "total_price": item["total_price"],
            "origin_country": item["origin_country"], "hawb": canonical["logistics"]["hawb_number"],
            "forwarder": canonical["international_forwarder"], "po": item["po_number"],
            "pieces": item["quantity"], "net": item["net_weight"], "gross": item["gross_weight"],
            "date": declaration_date,
        }
        for field, value in values.items():
            sheet.cell(output_row, columns[field]).value = value
        sheet.cell(output_row, 1).value = None
        if remark_column:
            sheet.cell(output_row, remark_column).value = "进境"
        for field_name, column in declaration_columns.items():
            if column:
                sheet.cell(output_row, column).value = item["declaration"][field_name]
    target = unique_download_path(f"{source.stem}_进境更新", source.suffix)
    workbook.save(target)
    workbook.close()
    return to_virtual_artifact_path(target)


def _generate_invoice_packing(canonical: Mapping[str, Any]) -> str:
    workbook = load_workbook(_INVOICE_TEMPLATE, data_only=False)
    customs = workbook["Customs invoice"]
    packing = workbook["Packing List"]
    now = datetime.now()
    customs["F2"] = now.date()
    packing["F3"] = now.date()
    customs_total = _prepare_detail_rows(customs, 32, 39, len(canonical["items"]), 11)
    packing_total = _prepare_detail_rows(packing, 37, 44, len(canonical["items"]), 11)
    currency = canonical["items"][0]["currency"] or "需确认"
    total_amount = Decimal("0")
    for index, item in enumerate(canonical["items"]):
        row = 32 + index
        customs.cell(row, 1).value = index + 1
        customs.cell(row, 2).value = item["normalized_product_id"]
        customs.cell(row, 3).value = item["description"] or "需确认：描述"
        customs.cell(row, 6).value = item["quantity"]
        customs.cell(row, 8).value = item["raw_country"] or "需确认：原产国"
        customs.cell(row, 9).value = item["unit_price"]
        customs.cell(row, 11).value = item["total_price"]
        total_amount += _decimal(item["total_price"]) or Decimal("0")

        row = 37 + index
        packing.cell(row, 1).value = index + 1
        packing.cell(row, 2).value = item["normalized_product_id"]
        packing.cell(row, 3).value = item["description"] or "需确认：描述"
        packing.cell(row, 6).value = item["quantity"]
        packing.cell(row, 7).value = item["net_weight"]
        packing.cell(row, 10).value = item["gross_weight"]
    customs.cell(25, 9).value = f"( {currency})"
    customs.cell(25, 11).value = f"({currency})"
    customs.cell(customs_total, 4).value = "Total Amount:"
    customs.cell(customs_total, 9).value = currency
    customs.cell(customs_total, 11).value = _number(total_amount)
    packing.cell(packing_total, 4).value = "TOTAL PCS:"
    packing.cell(packing_total, 6).value = _number(_sum_decimal(item["quantity"] for item in canonical["items"]))
    packing.cell(packing_total + 1, 4).value = "TOTAL NET WEIGHT:"
    packing.cell(packing_total + 1, 6).value = canonical["logistics"]["net_weight"]
    packing.cell(packing_total + 2, 4).value = "TOTAL GROSS WEIGHT: "
    packing.cell(packing_total + 2, 6).value = canonical["logistics"]["gross_weight"]
    target = unique_download_path("invoice_packing进境", ".xlsx")
    workbook.save(target)
    workbook.close()
    return to_virtual_artifact_path(target)


def _generate_bonded_checklist(
    canonical: Mapping[str, Any],
    units: Mapping[str, tuple[str, str, str]],
) -> str:
    workbook = load_workbook(_BONDED_TEMPLATE, data_only=False)
    header = workbook["表头"]
    body = workbook["表体"]
    header["M2"] = f"{datetime.now():%Y%m%d}"
    header["S2"] = canonical["logistics"]["shipper_country_name"]
    header["T2"] = "2244" if canonical["customs_mode"] == "快件" else "2233"
    _prepare_bonded_rows(body, len(canonical["items"]))
    currency_names = {"USD": "美元", "CNY": "人民币", "RMB": "人民币", "EUR": "欧元"}
    for index, item in enumerate(canonical["items"]):
        row = 2 + index
        declaration = item["declaration"]
        declaration_unit, legal_unit, legal_second = units[item["normalized_product_id"]]
        values = {
            2: index + 1, 4: index + 1, 6: item["normalized_product_id"],
            7: declaration["hs_code"], 8: declaration["chinese_name"], 9: item["description"],
            10: declaration_unit, 11: item["quantity"], 12: legal_unit, 13: item["quantity"],
            14: legal_second, 15: item["net_weight"], 16: item["unit_price"],
            17: item["total_price"], 18: item["origin_country"],
            19: currency_names.get(item["currency"], item["currency"]),
            20: item["gross_weight"], 21: item["net_weight"],
            32: "1" if item["raw_country"] == "US" else "2",
        }
        for column, value in values.items():
            body.cell(row, column).value = value
    target = unique_download_path("核注清单导入模板_进境", ".xlsx")
    workbook.save(target)
    workbook.close()
    return to_virtual_artifact_path(target)


def _oracle_units(items: Sequence[Mapping[str, Any]]) -> tuple[dict[str, tuple[str, str, str]], list[dict[str, Any]]]:
    fallback = ("需确认：申报计量单位", "需确认：法定计量单位", "需确认：法定第二单位")
    result = {item["normalized_product_id"]: fallback for item in items}
    checks: list[dict[str, Any]] = []
    dsn, username, password = (os.getenv(name) for name in ("ORACLE_DSN", "ORACLE_USERNAME", "ORACLE_PASSWORD"))
    if not all((dsn, username, password)):
        return result, [{"field": "oracle_units", "reason": "Oracle 配置缺失"}]
    try:
        _init_oracle_client(os.getenv("ORACLE_CLIENT_LIB_DIR"))
        timeout = float(os.getenv("ORACLE_TIMEOUT_SECONDS") or "30")
        import oracledb

        with oracledb.connect(user=username, password=password, dsn=dsn, tcp_connect_timeout=timeout) as connection:
            if hasattr(connection, "call_timeout"):
                connection.call_timeout = int(timeout * 1000)
            with connection.cursor() as cursor:
                for item in items:
                    product_id = item["normalized_product_id"]
                    for candidate in _unit_candidates(item):
                        cursor.execute(_ORACLE_SQL, product_id=candidate)
                        row = cursor.fetchone()
                        if row:
                            result[product_id] = tuple(str(value).strip() if value not in (None, "") else fallback[index] for index, value in enumerate(row[:3]))
                            break
                    if result[product_id] == fallback:
                        checks.append({"field": f"items.{product_id}.oracle_units", "reason": "Oracle 未找到单位"})
    except Exception:
        return result, [{"field": "oracle_units", "reason": "Oracle 查询失败"}]
    return result, checks


def _init_oracle_client(lib_dir: str | None) -> None:
    global _ORACLE_CLIENT_INITIALIZED
    if not lib_dir or _ORACLE_CLIENT_INITIALIZED:
        return
    import oracledb

    oracledb.init_oracle_client(lib_dir=lib_dir)
    _ORACLE_CLIENT_INITIALIZED = True


def _unit_candidates(item: Mapping[str, Any]) -> list[str]:
    result: list[str] = []
    for value in (item["normalized_product_id"], item["product_id"], item["product_id_raw"]):
        for candidate in (clean_product_id(value), str(value).strip()):
            if candidate and candidate not in result:
                result.append(candidate)
    return result


def _prepare_detail_rows(sheet: Any, start: int, total: int, count: int, max_column: int) -> int:
    if count <= 1:
        return total
    sheet.insert_rows(start + 1, amount=count - 1)
    for row in range(start + 1, start + count):
        _copy_style_row(sheet, start, row, max_column, copy_values=False)
    return total + count - 1


def _prepare_bonded_rows(sheet: Any, count: int) -> None:
    if count <= 1:
        return
    sheet.insert_rows(3, amount=count - 1)
    for row in range(3, 2 + count):
        _copy_style_row(sheet, 2, row, 32, copy_values=True)


def _copy_sheet_row(sheet: Any, source_row: int, target_row: int, max_column: int) -> None:
    source_row = max(2, min(source_row, sheet.max_row))
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        value = source.value
        if isinstance(value, str) and value.startswith("="):
            try:
                value = Translator(value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                pass
        target.value = value
        _copy_cell_style(source, target)


def _copy_style_row(sheet: Any, source_row: int, target_row: int, max_column: int, *, copy_values: bool) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target.value = source.value if copy_values else None
        _copy_cell_style(source, target)


def _copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.protection = copy.copy(source.protection)
    target.number_format = source.number_format


def _headers(sheet: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        header = re.sub(r"\s+", "", str(sheet.cell(1, column).value or "")).lower()
        if header and header not in result:
            result[header] = column
    return result


def _optional_column(headers: Mapping[str, int], aliases: Sequence[str]) -> int | None:
    for alias in aliases:
        column = headers.get(re.sub(r"\s+", "", alias).lower())
        if column:
            return column
    return None


def _required_column(headers: Mapping[str, int], aliases: Sequence[str], field: str) -> int:
    column = _optional_column(headers, aliases)
    if column is None:
        raise ValueError(f"tracking workbook missing required column: {field}")
    return column


def _last_data_row(sheet: Any) -> int:
    for row in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row, column).value not in (None, "") for column in range(1, sheet.max_column + 1)):
            return row
    return 2


def _empty_declaration() -> dict[str, Any]:
    return {field: f"需确认：{field}" for field in DECLARATION_FIELDS}


def _needs_c(source_artifact: str, reasons: Sequence[str]) -> dict[str, Any]:
    return {"status": "needs_c", "source_artifact": source_artifact, "reasons": list(reasons)}


def _normalize_forwarder(value: Any) -> str | None:
    text = str(value or "").strip().upper()
    return text if text in VALID_FORWARDERS else None


def clean_product_id(value: Any) -> str:
    return re.sub(r"[\s-]+", "", str(value or "").strip())


def normalize_product_id(value: Any) -> str:
    cleaned = clean_product_id(value)
    return cleaned[-12:] if len(cleaned) > 12 else cleaned


def _comparison_token(field: str, value: Any) -> str | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if field in {"pieces", "quantity", "gross_weight", "unit_price", "total_price"}:
        return _decimal_token(value)
    if field in {"shipper_country", "raw_country"}:
        return _country_code(value) or str(value).strip().casefold()
    if field == "currency":
        text = re.sub(r"[\s._-]+", " ", str(value).strip().upper())
        return {"$": "USD", "US DOLLAR": "USD", "US DOLLARS": "USD", "RMB": "CNY", "人民币": "CNY"}.get(text, text)
    if field == "product_id_raw":
        return normalize_product_id(value)
    if field == "po_number":
        return re.sub(r"\s+", "", str(value))
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def _canonical_value(field: str, value: Any) -> Any:
    if field in {"pieces", "quantity"}:
        number = _positive_decimal(value)
        if number is None or (field == "pieces" and number != number.to_integral_value()):
            return None
        return int(number) if number == number.to_integral_value() else _number(number)
    if field == "gross_weight":
        return _number(_positive_decimal(value))
    if field in {"unit_price", "total_price"}:
        return _number(_decimal(value))
    if field in {"shipper_country", "raw_country"}:
        return _country_code(value) or str(value).strip()
    if field == "currency":
        return _comparison_token(field, value)
    return value.strip() if isinstance(value, str) else value


def _country_code(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"[\s,./()_-]+", " ", str(value).strip().upper()).strip()
    match = re.search(r"\b([A-Z]{2})\b", text)
    if match and match.group(1) in COUNTRY_NAMES:
        return match.group(1)
    return COUNTRY_ALIASES.get(text)


def _decimal(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, (int, float)):
        try:
            result = Decimal(str(value))
            return result if result.is_finite() else None
        except InvalidOperation:
            return None
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", str(value).strip())
    if not match:
        return None
    try:
        result = Decimal(match.group(0).replace(",", ""))
        return result if result.is_finite() else None
    except InvalidOperation:
        return None


def _positive_decimal(value: Any) -> Decimal | None:
    number = _decimal(value)
    return number if number is not None and number > 0 else None


def _decimal_token(value: Any) -> str | None:
    number = _decimal(value)
    return format(number.normalize(), "f") if number is not None else None


def _number(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    return int(value) if value == value.to_integral_value() else float(value)


def _sum_decimal(values: Sequence[Any] | Any, *, allow_missing: bool = False) -> Decimal | None:
    result = Decimal("0")
    found = False
    for value in values:
        number = _decimal(value)
        if number is None:
            if allow_missing:
                continue
            return None
        found = True
        result += number
    return result if found else None


def _first_nonempty(values: Any) -> Any:
    return next((value for value in values if value is not None and (not isinstance(value, str) or value.strip())), None)


def _consistent_or_first(
    group: Sequence[Mapping[str, Any]],
    field: str,
    manual_checks: list[dict[str, Any]],
    product_id: str,
) -> Any:
    values = [row[field] for row in group if row[field] is not None]
    tokens = {_comparison_token(field, value) for value in values}
    if len(tokens) > 1:
        _manual(manual_checks, f"items.{product_id}.{field}", "同料号多行值不一致，采用首个值")
    return values[0] if values else None


def _po_numbers(values: Any) -> str | None:
    result: list[str] = []
    for value in values:
        if value is None:
            continue
        for candidate in str(value).split("/"):
            candidate = re.sub(r"\s+", "", candidate)
            if candidate and candidate not in result:
                result.append(candidate)
    return "/".join(result) or None


def _manual(checks: list[dict[str, Any]], field: str, reason: str) -> None:
    if not any(check.get("field") == field and check.get("reason") == reason for check in checks):
        checks.append({"field": field, "reason": reason})
